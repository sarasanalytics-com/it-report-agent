#!/usr/bin/env python3
"""
Sanity-check the downloaded Excel files before Claude processes them.

Verifies:
  1. Files exist and are non-empty.
  2. Expected sheets are present.
  3. Required columns exist in each sheet.

Exit code 0 = OK, 1 = validation failure.
"""

import pathlib
import sys

import openpyxl

DATA_DIR = pathlib.Path(__file__).resolve().parent.parent / "data"

# ---------------------------------------------------------------------------
# Expected structure — update these to match your actual Excel layouts
# ---------------------------------------------------------------------------
EXPECTED = {
    "asset_inventory.xlsx": {
        # sheet_name -> list of required columns (case-insensitive match)
        "Assets": [
            "Asset ID",
            "Asset Type",
            "Make",
            "Model",
            "Serial Number",
            "Purchase Date",
            "Assigned To",
            "Status",
        ],
    },
    "spend_tracker.xlsx": {
        "Laptop Procurement": [
            "Vendor",
            "Item",
            "Quantity",
            "Unit Price",
            "Total",
            "Purchase Date",
        ],
        "App Subscriptions": [
            "App Name",
            "Vendor",
            "Annual Cost",
            "Renewal Date",
            "Status",
        ],
    },
}

# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def normalize(name: str) -> str:
    return name.strip().lower()


def validate_file(filename: str, sheet_specs: dict) -> list[str]:
    """Return a list of error messages (empty = valid)."""
    errors: list[str] = []
    path = DATA_DIR / filename

    if not path.exists():
        return [f"{filename}: file not found in {DATA_DIR}"]
    if path.stat().st_size == 0:
        return [f"{filename}: file is empty (0 bytes)"]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    actual_sheets = {normalize(s): s for s in wb.sheetnames}

    for expected_sheet, required_cols in sheet_specs.items():
        key = normalize(expected_sheet)
        if key not in actual_sheets:
            errors.append(f"{filename}: missing sheet '{expected_sheet}' (found: {wb.sheetnames})")
            continue

        ws = wb[actual_sheets[key]]
        # Read header row (row 1)
        header_row = [normalize(str(c.value or "")) for c in next(ws.iter_rows(min_row=1, max_row=1))]

        for col in required_cols:
            if normalize(col) not in header_row:
                errors.append(f"{filename} → {expected_sheet}: missing column '{col}'")

    wb.close()
    return errors


def main() -> None:
    all_errors: list[str] = []

    for filename, sheet_specs in EXPECTED.items():
        print(f"Validating {filename} …")
        errs = validate_file(filename, sheet_specs)
        if errs:
            for e in errs:
                print(f"  ✗ {e}", file=sys.stderr)
            all_errors.extend(errs)
        else:
            print(f"  ✓ {filename} OK")

    if all_errors:
        print(f"\nValidation failed with {len(all_errors)} error(s).", file=sys.stderr)
        sys.exit(1)

    print("\nAll files validated successfully.")


if __name__ == "__main__":
    main()
