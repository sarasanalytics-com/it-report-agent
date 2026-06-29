#!/usr/bin/env python3
"""Resolve live Excel formulas to values after download.

Some source workbooks (e.g. the spend tracker) store the monthly totals as
*live formulas* whose computed result is NOT cached in the .xlsx. openpyxl can
only read cached values, so those cells come back blank and the report
under-counts. This module recalculates such formulas with the `formulas`
library and writes the results back into the ORIGINAL workbook — preserving its
exact sheet names, structure and already-correct cached values.

It is deliberately conservative: it only fills a formula cell when that cell's
cached value is currently missing, so cells that already carry a value are never
touched. Best-effort — on any error the file is left exactly as downloaded.
"""

from __future__ import annotations

import os
import glob
import shutil
import logging
import tempfile
import datetime as dt

import openpyxl

log = logging.getLogger("recalc")

# Excel error strings the recalc engine may emit — never inject these.
_ERROR_VALUES = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!",
                 "#NUM!", "#SPILL!", "#CALC!"}


def _is_injectable(v) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float, dt.datetime, dt.date)):
        return True
    if isinstance(v, str):
        return v.strip() != "" and v.strip().upper() not in _ERROR_VALUES
    return False


def recalc_uncached_formulas(path: str) -> int:
    """Fill in formula cells whose cached value is missing, in place. Returns the
    number of cells filled (0 if nothing needed doing). Never raises."""
    try:
        # Which formula cells are currently blank (no cached value)?
        cached = openpyxl.load_workbook(path, data_only=True)
        formula_wb = openpyxl.load_workbook(path)  # formulas + structure
        missing = []  # (sheet_title, coord)
        for ws in formula_wb.worksheets:
            cws = cached[ws.title] if ws.title in cached.sheetnames else None
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cv = cws[cell.coordinate].value if cws is not None else None
                        if cv is None:
                            missing.append((ws.title, cell.coordinate))
        if not missing:
            return 0  # every formula already has a cached value — nothing to do

        # Compute everything with the formulas engine into a temp file.
        import formulas  # heavy import; only when actually needed
        tmp = tempfile.mkdtemp(prefix="recalc_")
        try:
            model = formulas.ExcelModel().loads(path).finish()
            model.calculate()
            model.write(dirpath=tmp)
            outs = glob.glob(os.path.join(tmp, "*"))
            if not outs:
                return 0
            computed_wb = openpyxl.load_workbook(outs[0], data_only=True)
            # The engine upper-cases sheet titles; match case-insensitively.
            by_sheet = {ws.title.upper(): ws for ws in computed_wb.worksheets}

            filled = 0
            for title, coord in missing:
                cws = by_sheet.get(title.upper())
                if cws is None:
                    continue
                val = cws[coord].value
                if _is_injectable(val):
                    formula_wb[title][coord].value = val
                    filled += 1
            if filled:
                formula_wb.save(path)
            return filled
        finally:
            shutil.rmtree(tmp, ignore_errors=True)
    except Exception as exc:  # noqa: BLE001 - recalc must never break the fetch
        log.warning("recalc skipped for %s (%s)", os.path.basename(path), exc)
        return 0


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO)
    for p in sys.argv[1:]:
        n = recalc_uncached_formulas(p)
        print(f"{p}: filled {n} formula cell(s)")
