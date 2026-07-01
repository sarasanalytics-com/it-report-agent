#!/usr/bin/env python3
"""Print the schema (sheet names + column headers) of the downloaded IT
workbooks, so the column mapping can be verified against the real files.

Privacy: prints only the detected HEADER row of each sheet (column names) and a
row count — never data rows — so no employee/PII values are logged.

Run after scripts/fetch-excel.py has populated data/.
"""

import pathlib

import openpyxl

DATA_DIR = pathlib.Path(__file__).resolve().parent.parent / "data"

FILES = [
    "asset_inventory.xlsx",
    "spend_tracker.xlsx",
    "procurement_plan.xlsx",
    "joiners_info.xlsx",
    "vendor_payments.xlsx",
    "app_spend_source.xlsx",
]

# For these files, also print per-column numeric COUNT + SUM (aggregates only,
# no individual values) so we can locate money columns and confirm the amounts
# are real numbers rather than unresolved links.
VALUE_PROFILE_FILES = {"app_spend_source.xlsx"}


def _header_row(ws):
    """Pick the most header-like row in the first 6 rows (most non-empty cells)."""
    best_idx, best_cells, best_count = None, [], -1
    for i, raw in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        cells = [str(c).strip() for c in raw if c not in (None, "")]
        if len(cells) > best_count:
            best_idx, best_cells, best_count = i, cells, len(cells)
    return best_idx, best_cells


def main() -> None:
    for fname in FILES:
        path = DATA_DIR / fname
        print(f"\n{'=' * 70}\n# {fname}")
        if not path.exists():
            print("  (not downloaded)")
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            print(f"  ERROR opening: {exc}")
            continue
        print(f"  sheets: {wb.sheetnames}")
        profile = fname in VALUE_PROFILE_FILES
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            hdr_idx, headers = _header_row(ws)
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            print(f"\n  -- sheet: {sheet!r}  (~{len(rows)} rows, header looks like row {hdr_idx})")
            print(f"     columns: {headers}")
            if profile and rows:
                # Per-column numeric count + sum (aggregates only — no cell values).
                hdr = rows[hdr_idx - 1] if hdr_idx and hdr_idx <= len(rows) else rows[0]
                width = max((len(r) for r in rows), default=0)
                print("     numeric profile (col: header → count, sum):")
                for c in range(width):
                    nums = [r[c] for r in rows[hdr_idx:] if c < len(r)
                            and isinstance(r[c], (int, float)) and not isinstance(r[c], bool)]
                    if nums:
                        h = hdr[c] if c < len(hdr) and hdr[c] is not None else f"col{c}"
                        print(f"       [{c}] {str(h)[:30]!r} → {len(nums)} nums, sum={round(sum(nums), 2)}")
        wb.close()


if __name__ == "__main__":
    main()
