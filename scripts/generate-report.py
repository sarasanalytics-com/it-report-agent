#!/usr/bin/env python3
"""
Generate IT weekly or monthly reports from Excel data.

Usage:
    python generate-report.py weekly
    python generate-report.py monthly

Reads Excel files from data/ and writes:
    output/slack-summary.md
    output/full-report.md
"""

from __future__ import annotations

import json
import os
import re
import sys
import pathlib
import datetime as dt
from datetime import timedelta
from collections import defaultdict
from typing import Optional

import openpyxl

DATA_DIR = pathlib.Path(__file__).resolve().parent.parent / "data"
OUTPUT_DIR = pathlib.Path(__file__).resolve().parent.parent / "output"
SNAPSHOT_DIR = pathlib.Path(__file__).resolve().parent.parent / "snapshots"
TODAY = dt.datetime.now().date()
AGE_THRESHOLD_DAYS = int(3.5 * 365)  # 1277 days

# Reporting currency is USD ($). Laptop procurement and budget figures in the
# source spreadsheets are recorded in INR, so they are converted to USD for the
# report. Override the rate with the INR_TO_USD_RATE env var when it drifts.
DEFAULT_INR_TO_USD_RATE = 0.0117  # ≈ ₹85.5 / $1


def _load_inr_rate() -> float:
    """Read INR_TO_USD_RATE from the environment, tolerating an unset or blank
    value (e.g. a workflow secret that hasn't been configured) by falling back
    to the default."""
    raw = os.environ.get("INR_TO_USD_RATE", "").strip()
    if raw:
        try:
            rate = float(raw)
            if rate > 0:
                return rate
        except ValueError:
            print(f"Warning: invalid INR_TO_USD_RATE={raw!r}; using default "
                  f"{DEFAULT_INR_TO_USD_RATE}", file=sys.stderr)
    return DEFAULT_INR_TO_USD_RATE


INR_TO_USD_RATE = _load_inr_rate()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def is_truly_assigned(row: dict) -> bool:
    """True if the row represents an actively assigned laptop.
    Excludes rows where Employee ID is blank or marked "In Stock" /
    "Hand over" / similar non-assignment placeholders."""
    emp_id = row.get("Employee ID")
    if emp_id is None:
        return False
    emp_id_str = str(emp_id).strip().lower()
    if not emp_id_str:
        return False
    # Common non-assignment markers in the Employee ID column
    non_assignment = ("in stock", "instock", "hand over", "handover", "stock")
    if any(marker in emp_id_str for marker in non_assignment):
        return False
    return True


def parse_date(val) -> Optional[dt.date]:
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        # Month-column headers like 'Jun-26', 'Jun-2026', 'June 2026', '2026-06'.
        for fmt in ("%b-%y", "%b-%Y", "%B-%y", "%B-%Y", "%b %y", "%b %Y",
                    "%B %y", "%B %Y", "%Y-%m"):
            try:
                return dt.datetime.strptime(s, fmt).date().replace(day=1)
            except ValueError:
                continue
    return None


def age_years(dt: datetime.date) -> float:
    return (TODAY - dt).days / 365.25


def fmt_usd(amount) -> str:
    try:
        n = float(amount)
    except (TypeError, ValueError):
        return "N/A"
    if n >= 1_000_000:
        return f"${n/1_000_000:,.2f}M"
    if n >= 1_000:
        return f"${n:,.2f}"
    return f"${n:,.2f}"


def fmt_inr(amount) -> str:
    try:
        n = float(amount)
    except (TypeError, ValueError):
        return "N/A"
    if n >= 10_000_000:
        return f"₹{n/10_000_000:,.2f} Cr"
    if n >= 100_000:
        return f"₹{n/100_000:,.2f} L"
    return f"₹{n:,.0f}"


def fmt_inr_full(amount) -> str:
    """Exact INR amount with thousands separators, e.g. ₹188,850 or
    ₹165,787.64 (no lakh/crore abbreviation — invoice figures stay precise)."""
    try:
        n = float(amount)
    except (TypeError, ValueError):
        return "N/A"
    if abs(n - round(n)) < 0.005:
        return f"₹{int(round(n)):,}"
    return f"₹{n:,.2f}"


def inr_to_usd(amount) -> float:
    """Convert an INR amount to USD using the configured rate. Returns 0.0 for
    non-numeric input."""
    try:
        return float(amount) * INR_TO_USD_RATE
    except (TypeError, ValueError):
        return 0.0


def fmt_usd_from_inr(amount) -> str:
    """Format an INR-denominated amount as USD."""
    return fmt_usd(inr_to_usd(amount))


def _to_number(val) -> Optional[float]:
    """Parse a spreadsheet cell into a float, tolerating numbers stored as text
    with currency symbols / thousands separators (e.g. "$1,234.50", "1,234").
    Returns None when the value isn't numeric."""
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        for token in ("$", "₹", "USD", "INR", "Rs.", "Rs", ","):
            s = s.replace(token, "")
        s = s.strip()
        if not s:
            return None
        # Handle parenthesised negatives e.g. (1,234)
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except ValueError:
            return None
    return None


def read_sheet(wb, sheet_name: str, header_row: int = 1) -> list[dict]:
    """Read a sheet into a list of dicts. Returns [] if sheet missing."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def read_sheet_auto(wb, sheet_name: str) -> list[dict]:
    """Like read_sheet but auto-detects the header row (the most-populated of the
    first 6 rows). Use for hand-made sheets that may have a title/blank row first."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    best_idx, best_count = 1, -1
    for i, raw in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        n = sum(1 for c in raw if c not in (None, ""))
        if n > best_count:
            best_idx, best_count = i, n
    return read_sheet(wb, sheet_name, header_row=best_idx)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_data() -> dict:
    asset_wb = openpyxl.load_workbook(DATA_DIR / "asset_inventory.xlsx", read_only=True, data_only=True)
    spend_wb = openpyxl.load_workbook(DATA_DIR / "spend_tracker.xlsx", read_only=True, data_only=True)
    proc_wb = openpyxl.load_workbook(DATA_DIR / "procurement_plan.xlsx", read_only=True, data_only=True)
    join_wb = openpyxl.load_workbook(DATA_DIR / "joiners_info.xlsx", read_only=True, data_only=True)

    data = {
        "assigned": read_sheet(asset_wb, "Laptop Assigned"),
        "in_stock": read_sheet(asset_wb, "Laptop in stock"),
        "backup": read_sheet(asset_wb, "Backup Laptops 3years old"),
        "history": read_sheet(asset_wb, "Assset History"),
        "returned": read_sheet(asset_wb, "Laptop Returned"),
        "pending_returns": read_sheet(asset_wb, "Laptop yet to Return"),
        "purchased": read_sheet(asset_wb, "New Laptops purchased "),
        "sold": read_sheet(asset_wb, "Laptops sold "),
        "mouse": read_sheet(asset_wb, "Mouse"),
        "headset": read_sheet(asset_wb, "Headset"),
        "keyboard": read_sheet(asset_wb, "Keyboard"),
        "charger": read_sheet(asset_wb, "Charger"),
        "docking": read_sheet(asset_wb, "Docking station"),
        "monitor": read_sheet(asset_wb, "Monitor"),
        "other_stock": read_sheet(asset_wb, "Other Assets Instock"),
        # Main subscriptions sheet + the extra 'Linkdin Growth Team' apps so the
        # software total/inventory is complete.
        "spend": (read_sheet(spend_wb, "Sheet1")
                  + read_sheet(spend_wb, "Linkdin Growth Team")),
        "joinings": read_sheet(join_wb, "Joinings"),
        "checklist": read_sheet(join_wb, "Joining checklist"),
        "proc_plan": read_sheet(proc_wb, "Laptop procurement plan", header_row=2),
        "actual_spend": read_sheet(proc_wb, "Actual Spends", header_row=3),
        "configuration": read_sheet(proc_wb, "Configuration"),
    }

    # IT helpdesk tickets. Preferred source is data/it_issues.xlsx, produced by
    # scripts/fetch-issues.py from the ClickUp IT ticket list. Falls back to an
    # optional "IT Issues" sheet inside the asset workbook, and finally to empty
    # (the report then shows a placeholder).
    data["it_issues"] = _load_it_issues(asset_wb)

    # Vendor payments (optional). Downloaded from SharePoint to
    # data/vendor_payments.xlsx by scripts/fetch-excel.py. First sheet is read.
    data["vendor"] = _load_vendor_payments()

    # Laptop delivery lead-times + vendor payment terms (optional budget tables).
    data["delivery"] = _load_delivery_timelines(proc_wb)
    data["payment_terms"] = _load_payment_terms(proc_wb)

    # Unplanned / ad-hoc spends — a dedicated sheet the IT owner maintains
    # (auto-detected by name). Read as-is so we never reinterpret the figures.
    data["unplanned"] = _load_unplanned_spends([proc_wb, spend_wb, asset_wb, join_wb])

    for wb in (asset_wb, spend_wb, proc_wb, join_wb):
        wb.close()

    # A 'ready stock' row only counts as a real laptop if it carries some
    # identifier (make/model/serial/tag). Rows with just stray cells (e.g. a
    # lone 'Condition') are junk and shouldn't inflate the Stock Ready count.
    _ID_FIELDS = ("Laptop Make", "Laptop Model", "Laptop Serial Number",
                  "Serial Number", "Serial No", "Laptop Asset Tag")
    kept = [r for r in data["in_stock"]
            if any(str(r.get(k, "") or "").strip() for k in _ID_FIELDS)]
    dropped = len(data["in_stock"]) - len(kept)
    if dropped:
        print(f"  Excluded {dropped} in-stock row(s) with no identifier (blank/junk).",
              file=sys.stderr)
    data["in_stock"] = kept

    # Surface any kept 'ready stock' row that has an identifier but no make/model
    # so the 'Unspecified — …' entries can be identified and completed at source.
    for row in data["in_stock"]:
        if not str(row.get("Laptop Make", "") or "").strip() and \
           not str(row.get("Laptop Model", "") or "").strip():
            populated = {k: v for k, v in row.items() if v not in (None, "")}
            print(f"  Note: in-stock laptop with no make/model → {populated}",
                  file=sys.stderr)

    return data


def _load_vendor_payments() -> list[dict]:
    """Load vendor payment rows from data/vendor_payments.xlsx.

    The sheet often has title rows above the real header, and the table may not
    be on the first tab, so scan every sheet for the header row (the one
    containing a 'Vendor'-ish column plus 'Amount'/'Status') and read from
    there. Returns the rows from the first sheet that yields data, else [].
    """
    path = DATA_DIR / "vendor_payments.xlsx"
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # corrupt/locked file shouldn't break the report
        print(f"  Warning: could not read vendor_payments.xlsx: {exc}", file=sys.stderr)
        return []

    def _is_header(cells: list[str]) -> bool:
        joined = " ".join(cells)
        return "vendor" in joined and ("amount" in joined or "status" in joined or "inv" in joined)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        scan = list(ws.iter_rows(min_row=1, max_row=25, values_only=True))
        hdr_idx = None
        for idx, raw in enumerate(scan):
            cells = [str(c).strip().lower() if c is not None else "" for c in raw]
            if _is_header(cells):
                hdr_idx = idx
                break
        if hdr_idx is None:
            continue
        headers = [str(c).strip() if c is not None else f"col{j}"
                   for j, c in enumerate(scan[hdr_idx])]
        rows = []
        for raw in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
            if all(c is None for c in raw):
                continue
            rows.append({headers[j]: raw[j] for j in range(min(len(headers), len(raw)))})
        if rows:
            print(f"  Vendor payments: sheet '{sheet_name}', header row {hdr_idx + 1}, "
                  f"{len(rows)} data row(s)")
            wb.close()
            return rows

    print("  Vendor payments: no recognizable table found in "
          f"{wb.sheetnames}", file=sys.stderr)
    wb.close()
    return []


_DELIVERY_META_COLS = ("device type", "processor", "department")


def _load_delivery_timelines(proc_wb) -> dict:
    """Find and read the laptop delivery lead-time matrix from the budget
    workbook. Auto-detects the sheet/header by looking for a 'Device Type'
    column. Returns {headers, rows} (empty if not present)."""
    try:
        # Prefer a sheet whose name signals delivery timelines; the "Configuration"
        # sheet also has a "Device Type" column, so for other sheets we also
        # require a "Processor" column to avoid matching it.
        named = [s for s in proc_wb.sheetnames
                 if any(k in s.lower() for k in ("deliver", "timeline", "lead time", "lead-time"))]
        ordered = named + [s for s in proc_wb.sheetnames if s not in named]
        for sheet in ordered:
            ws = proc_wb[sheet]
            scan = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
            hdr_idx = None
            for i, raw in enumerate(scan):
                cells = [str(c).strip().lower() if c is not None else "" for c in raw]
                has_device = any("device type" in c for c in cells)
                has_proc = any(c == "processor" for c in cells)
                if has_device and (sheet in named or has_proc):
                    hdr_idx = i
                    break
            if hdr_idx is None:
                continue
            headers = [str(c).strip() if c is not None else "" for c in scan[hdr_idx]]
            rows = []
            for raw in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
                if all(c is None for c in raw):
                    if rows:  # blank row ends the table (e.g. payment-terms below)
                        break
                    continue
                device = str(raw[0] if raw else "" or "").strip()
                if "payment term" in device.lower():  # another table starts here
                    break
                rec = {headers[j]: raw[j] for j in range(min(len(headers), len(raw)))}
                if str(rec.get("Device Type", "") or "").strip():
                    rows.append(rec)
            if rows:
                print(f"  Delivery timelines: sheet '{sheet}', header row {hdr_idx + 1}, "
                      f"{len(rows)} device row(s)")
                return {"headers": headers, "rows": rows}
    except Exception as exc:  # never break the report over this optional table
        print(f"  Warning: could not read delivery timelines: {exc}", file=sys.stderr)
    return {"headers": [], "rows": []}


def _load_payment_terms(proc_wb) -> dict:
    """Vendor → payment terms, from a 'Payment terms' row in the budget workbook
    (header cell 'Payment terms' followed by vendor columns; the next row holds
    each vendor's terms). Returns {vendor: terms} (empty if not present)."""
    try:
        for sheet in proc_wb.sheetnames:
            ws = proc_wb[sheet]
            rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))
            for i, raw in enumerate(rows):
                cells = [str(c).strip() if c is not None else "" for c in raw]
                lower = [c.lower() for c in cells]
                pt_idx = next((j for j, c in enumerate(lower) if "payment term" in c), None)
                if pt_idx is None:
                    continue
                vendor_cols = {j: cells[j] for j in range(len(cells))
                               if j != pt_idx and cells[j]}
                if not vendor_cols:
                    continue
                for raw2 in rows[i + 1:]:
                    c2 = [str(c).strip() if c is not None else "" for c in raw2]
                    terms = {v: c2[j] for j, v in vendor_cols.items() if j < len(c2) and c2[j]}
                    if terms:
                        print(f"  Payment terms: sheet '{sheet}', {len(terms)} vendor(s)")
                        return terms
    except Exception as exc:
        print(f"  Warning: could not read payment terms: {exc}", file=sys.stderr)
    return {}


def get_payment_terms(data: dict) -> dict:
    """Cleaned {vendor: payment terms} mapping."""
    return {str(k).strip(): " ".join(str(v).split())
            for k, v in (data.get("payment_terms") or {}).items() if str(v).strip()}


def _load_unplanned_spends(wbs: list) -> dict:
    """Find a dedicated 'Unplanned' spends sheet (the IT owner maintains it for
    ad-hoc / off-budget purchases) and read it AS-IS, so its figures are never
    reinterpreted. Auto-detects any sheet whose name mentions 'unplan' (so
    'Unplaned'/'unplanned'/'un-planned' all hit) or 'unbudgeted'/'ad hoc'.

    Scans EVERY downloaded workbook in data/ — not just a fixed few — so it
    doesn't matter which file the tab lives in. Returns {sheet, headers, rows}."""
    def _is_unplanned(name: str) -> bool:
        n = name.strip().lower().replace("-", "").replace(" ", "")
        return ("unplan" in n or "unbudget" in n or "adhoc" in n
                or "outofbudget" in n)

    # Prefer the already-open workbooks (no re-read), then any other spreadsheet
    # in data/ we haven't covered (e.g. the vendor payments file).
    for wb in wbs:
        try:
            for sheet in wb.sheetnames:
                if not _is_unplanned(sheet):
                    continue
                rows = read_sheet_auto(wb, sheet)
                if not rows:
                    continue
                headers = list(rows[0].keys())
                print(f"  Unplanned spends: sheet '{sheet}', {len(rows)} row(s)")
                return {"sheet": sheet, "headers": headers, "rows": rows}
        except Exception as exc:  # noqa: BLE001
            print(f"  Warning: could not read unplanned spends: {exc}", file=sys.stderr)

    for path in sorted(DATA_DIR.glob("*.xlsx")):
        try:
            extra = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:  # noqa: BLE001
            continue
        try:
            for sheet in extra.sheetnames:
                if not _is_unplanned(sheet):
                    continue
                rows = read_sheet_auto(extra, sheet)
                if not rows:
                    continue
                headers = list(rows[0].keys())
                print(f"  Unplanned spends: sheet '{sheet}' in {path.name}, "
                      f"{len(rows)} row(s)")
                return {"sheet": sheet, "headers": headers, "rows": rows}
        finally:
            extra.close()
    return {"sheet": None, "headers": [], "rows": []}


def get_unplanned_spends(data: dict) -> dict:
    """Unplanned / ad-hoc spends, from EITHER a dedicated 'Unplanned' tab (read
    verbatim) OR 'UnPlaned'-labelled rows inside the 'Laptop procurement plan'
    sheet (the Department column), kept SEPARATE from the planned budget.

    Returns a dict whose 'mode' is one of:
      'sheet'      -> {mode, sheet, headers, rows, amount_col, total}
      'plan_rows'  -> {mode, sheet, items:[{category,model,quantity,amount_inr}], total}
      'none'       -> {mode}
    """
    # 1) Dedicated tab — reproduced verbatim (its own columns).
    raw = data.get("unplanned") or {}
    rows = raw.get("rows") or []
    headers = [h for h in (raw.get("headers") or [])
               if not str(h).startswith("col_")]  # drop blank-header spacer cols
    if rows:
        amount_col = None
        for h in headers:
            hl = str(h).lower()
            if any(k in hl for k in ("amount", "cost", "price", "spend", "value",
                                     "total", "inr", "₹")):
                amount_col = h
                break
        out_rows, total, any_amount = [], 0.0, False
        for r in rows:
            if not any(str(v).strip() for v in r.values() if v is not None):
                continue  # blank spacer row
            clean = {h: r.get(h) for h in headers}
            if amount_col is not None:
                n = _to_number(r.get(amount_col))
                if n is not None:
                    total += n
                    any_amount = True
            out_rows.append(clean)
        if out_rows:
            return {"mode": "sheet", "sheet": raw.get("sheet"), "headers": headers,
                    "rows": out_rows, "amount_col": amount_col,
                    "total": total if any_amount else None}

    # 2) 'UnPlaned' rows inside the Laptop procurement plan sheet.
    items, total, any_amount = [], 0.0, False
    for row in data.get("proc_plan", []):
        dept = str(row.get("Department", "") or "")
        if "unplan" not in dept.lower().replace("-", "").replace(" ", ""):
            continue
        model = str(row.get("Model", "") or "").strip()
        qty = _to_number(row.get("Quantity"))
        amt = _to_number(row.get("Total Price (INR)"))
        if not model and qty in (None, 0) and amt is None:
            continue  # bare section-header row, nothing to show
        items.append({"category": " ".join(dept.split()), "model": model,
                      "quantity": int(qty) if qty is not None else None,
                      "amount_inr": amt})
        if amt:
            total += amt
            any_amount = True
    if items:
        return {"mode": "plan_rows", "sheet": "Laptop procurement plan",
                "items": items, "total": total if any_amount else None}

    return {"mode": "none"}


def _leadtime_to_days(text) -> Optional[int]:
    """Worst-case lead time in days from text like '1 to 2 days', '1 week',
    '1 to 2 weeks', '3 days'. Returns None if not parseable."""
    s = str(text or "").lower()
    nums = [int(n) for n in re.findall(r"\d+", s)]
    if not nums:
        return None
    worst = max(nums)
    return worst * 7 if "week" in s else worst


def get_delivery_options(data: dict) -> list[dict]:
    """Structured delivery options per device: vendors with lead-time text/days
    and the fastest vendor."""
    d = data.get("delivery", {}) or {}
    headers = [h for h in d.get("headers", []) if h]
    vendor_cols = [h for h in headers if h.strip().lower() not in _DELIVERY_META_COLS]
    options = []
    for rec in d.get("rows", []):
        vendors = []
        for vc in vendor_cols:
            txt = str(rec.get(vc, "") or "").strip()
            if not txt:
                continue
            vendors.append({"vendor": vc.strip(), "text": txt, "days": _leadtime_to_days(txt)})
        timed = [v for v in vendors if v["days"] is not None]
        options.append({
            "device": str(rec.get("Device Type", "") or "").strip(),
            "processor": str(rec.get("Processor", "") or "").strip(),
            "departments": str(rec.get("Department", "") or rec.get("Department ", "") or "").strip(),
            "vendors": vendors,
            "fastest": min(timed, key=lambda v: v["days"]) if timed else None,
        })
    return options


def get_fastest_delivery_for_department(data: dict, department: str) -> Optional[dict]:
    """Best (fastest) delivery option for a given department, matched against the
    'Department' column of the delivery matrix. Returns the option dict or None."""
    dept = str(department or "").strip().lower()
    if not dept:
        return None
    best = None
    for opt in get_delivery_options(data):
        depts = opt["departments"].lower()
        if dept and (dept in depts or any(d.strip() and d.strip() in dept for d in depts.split(","))):
            if opt["fastest"] and (best is None or opt["fastest"]["days"] < best["fastest"]["days"]):
                best = opt
    return best


def _short_device(name) -> str:
    s = str(name or "").strip()
    for pre in ("Apple ", "Windows Laptop ", "Windows ", "Laptop "):
        if s.startswith(pre):
            s = s[len(pre):]
    return s.strip() or str(name or "").strip()


def get_vendor_delivery_matrix(data: dict) -> dict:
    """Vendor-centric pivot: each vendor's delivery lead time per device plus its
    payment terms. Returns {devices: [short names], rows: [{vendor, cells, terms,
    min_days}]} ordered fastest-first."""
    options = get_delivery_options(data)
    terms = get_payment_terms(data)
    devices = [_short_device(o["device"]) for o in options]
    vendors, cell, days = [], {}, {}
    for o in options:
        short = _short_device(o["device"])
        for v in o["vendors"]:
            if v["vendor"] not in vendors:
                vendors.append(v["vendor"])
            cell[(v["vendor"], short)] = v["text"]
            if v["days"] is not None:
                days.setdefault(v["vendor"], []).append(v["days"])
    rows = []
    for vendor in vendors:
        rows.append({
            "vendor": vendor,
            "cells": [cell.get((vendor, d), "—") for d in devices],
            "terms": terms.get(vendor, "—"),
            "min_days": min(days.get(vendor, [9999])),
        })
    rows.sort(key=lambda r: r["min_days"])
    return {"devices": devices, "rows": rows}


def _load_it_issues(asset_wb) -> list[dict]:
    """Load IT issues from data/it_issues.xlsx if present, else the asset
    workbook's optional 'IT Issues' sheet, else []."""
    issues_path = DATA_DIR / "it_issues.xlsx"
    if issues_path.exists():
        try:
            iwb = openpyxl.load_workbook(issues_path, read_only=True, data_only=True)
            rows = read_sheet(iwb, "IT Issues")
            iwb.close()
            return rows
        except Exception as exc:  # corrupt/locked file shouldn't break the report
            print(f"  Warning: could not read it_issues.xlsx: {exc}", file=sys.stderr)
    return read_sheet(asset_wb, "IT Issues")


# ---------------------------------------------------------------------------
# Analysis functions
# ---------------------------------------------------------------------------

def get_stock_summary(data: dict) -> dict:
    return {
        "Laptops (ready)": len(data["in_stock"]),
        "Laptops (3yr+ backup)": len(data["backup"]),
    }


def _os_label(value) -> str:
    """Normalise an Operating System string to a tidy bucket."""
    s = str(value or "").strip().lower()
    if not s:
        return "Unspecified"
    if "mac" in s or "osx" in s or "os x" in s:
        return "macOS"
    if "win" in s:
        return "Windows"
    if "ubuntu" in s or "linux" in s or "fedora" in s:
        return "Linux"
    if "chrome" in s:
        return "ChromeOS"
    return str(value).strip().title()


def _clean_processor(value) -> str:
    """Shorten verbose CPU strings to a readable chip name, e.g.
    '11th Gen Intel(R) Core(TM) i5-11320H @ 3.20GHz  3.19 GHz' -> 'i5-11320H',
    '13th Gen Intel(R) Core(TM) i7-1355U, 1700 MHz, 10 Cores' -> 'i7-1355U',
    'Ultra 7-155H' -> 'Ultra 7-155H', 'Apple M3 Pro' -> 'M3 Pro'."""
    s = str(value or "").strip()
    if not s or s.lower() == "none":
        return ""
    m = re.search(r"\bi[3579][\s-]?\d{3,5}\w*", s, re.I)
    if m:
        return m.group(0).replace(" ", "-")
    m = re.search(r"\bUltra\s*\d+[\s-]?\w*", s, re.I)
    if m:
        return " ".join(m.group(0).split())
    m = re.search(r"\bRyzen\s*\d+\s*\w*", s, re.I)
    if m:
        return " ".join(m.group(0).split())
    m = re.search(r"\bM[1-4](?:\s*(?:Pro|Max|Ultra))?\b", s)
    if m:
        return m.group(0)
    # Fallback: strip the common noise tokens and clock speeds.
    s = re.sub(r"\(R\)|\(TM\)|Intel|Core|\d+th\s*Gen|@.*|,.*|\d+(?:\.\d+)?\s*[GM]Hz.*"
               r"|\d+\s*Cores.*|\d+\s*Logical.*", " ", s, flags=re.I)
    return " ".join(s.split())[:24]


def get_stock_by_os(data: dict) -> dict:
    """Group ready stock laptops by OS. Returns {os_label: [config dicts]},
    ordered by count descending. Each config dict has make/model/ram/processor
    and a one-line 'config' string."""
    groups: dict[str, list] = defaultdict(list)
    for row in data["in_stock"]:
        make = str(row.get("Laptop Make", "") or "").strip()
        model = str(row.get("Laptop Model", "") or "").strip()
        ram = " ".join(str(row.get("RAM", "") or "").strip().split())
        proc = _clean_processor(row.get("Processor"))
        tag = str(row.get("Laptop Asset Tag", "") or "").strip()
        serial = str(row.get("Laptop Serial Number", "")
                     or row.get("Serial Number", "") or row.get("Serial No", "")
                     or "").strip()
        parts = [p for p in (f"{make} {model}".strip(), ram, proc)
                 if p and p.lower() not in ("none", "")]
        config = " · ".join(parts)
        if not config:
            # No specs — surface whatever identifies the unit so it isn't a
            # nameless "details not recorded" row.
            config = tag or (f"S/N {serial}" if serial else "") or "details not recorded"
        groups[_os_label(row.get("Operating System"))].append({
            "make": make, "model": model, "ram": ram, "processor": proc,
            "tag": tag, "serial": serial, "config": config,
        })
    return dict(sorted(groups.items(), key=lambda kv: len(kv[1]), reverse=True))


# Vendor-payment statuses that count as still-owed.
_VENDOR_PAID = ("paid", "completed", "complete", "done", "closed", "cancelled", "canceled")


def get_vendor_payments(data: dict) -> dict:
    """Pending vendor payments from the optional vendor sheet.

    Columns: Vendor Name, Inv. No, Date, Amount, Terms, Due Dt.,
    Overdue By(Days), Status. Amounts are kept in their native INR. Returns
    {connected, pending(list), total_inr, count}.
    """
    rows = data.get("vendor", [])

    def _get(row, *names):
        for n in names:
            for k, v in row.items():
                if str(k).strip().lower() == n.lower():
                    return v
        return None

    pending = []
    total = 0.0
    for row in rows:
        vendor = str(_get(row, "Vendor Name", "Vendor") or "").strip()
        status = str(_get(row, "Status") or "").strip()
        if not vendor and not status:
            continue
        if status.lower() in _VENDOR_PAID:
            continue
        amt_inr = _to_number(_get(row, "Amount")) or 0.0
        pending.append({
            "vendor": vendor,
            "invoice": str(_get(row, "Inv. No", "Invoice No", "Invoice") or "").strip(),
            "amount_inr": amt_inr,
            "due": parse_date(_get(row, "Due Dt.", "Due Date", "Due")),
            "overdue": str(_get(row, "Overdue By(Days)", "Overdue By", "Overdue") or "").strip(),
            "status": status,
        })
        total += amt_inr
    pending.sort(key=lambda x: x["due"] or dt.date.max)
    return {"connected": bool(rows), "pending": pending,
            "total_inr": total, "count": len(pending)}


def aging_action(a: dict) -> str:
    """What to do with an aging laptop — for the 'action' column."""
    if a["priority"] == "Critical":
        action = "Replace now (>4 yrs)"
    else:
        action = "Plan replacement (3.5–4 yrs)"
    we = a.get("warranty_end")
    if we and we < TODAY:
        action += f"; warranty expired {we.strftime('%b %Y')}"
    elif we and (we - TODAY).days <= 90:
        action += f"; warranty ends {we.strftime('%b %Y')}"
    return action


def get_recent_assignments(data: dict, days: int = 7) -> list[dict]:
    cutoff = TODAY - timedelta(days=days)
    results = []
    for row in data["history"]:
        dt = parse_date(row.get("Assigned Date"))
        if dt and cutoff <= dt <= TODAY:
            results.append(row)
    return results


def get_assignments_in_window(data: dict, start_days_ago: int, end_days_ago: int) -> list[dict]:
    """Assignments with date in [TODAY - start_days_ago, TODAY - end_days_ago]."""
    start = TODAY - timedelta(days=start_days_ago)
    end = TODAY - timedelta(days=end_days_ago)
    results = []
    for row in data["history"]:
        d = parse_date(row.get("Assigned Date"))
        if d and start <= d <= end:
            results.append(row)
    return results


def get_returns_in_window(data: dict, start_days_ago: int, end_days_ago: int) -> list[dict]:
    start = TODAY - timedelta(days=start_days_ago)
    end = TODAY - timedelta(days=end_days_ago)
    results = []
    for row in data["returned"]:
        d = parse_date(row.get("Returned Date"))
        if d and start <= d <= end:
            results.append(row)
    return results


def get_recent_returns(data: dict, days: int = 7) -> list[dict]:
    cutoff = TODAY - timedelta(days=days)
    results = []
    for row in data["returned"]:
        dt = parse_date(row.get("Returned Date"))
        if dt and cutoff <= dt <= TODAY:
            results.append(row)
    return results


def get_weekly_activity_comparison(data: dict) -> dict:
    """Compare this-week (last 7d) vs previous-week (7-14d ago) activity."""
    this_assigns = get_assignments_in_window(data, 7, 0)
    prev_assigns = get_assignments_in_window(data, 14, 8)
    this_repl = [a for a in this_assigns if str(a.get("New Joiner/Replacement", "")).lower() == "replacement"]
    prev_repl = [a for a in prev_assigns if str(a.get("New Joiner/Replacement", "")).lower() == "replacement"]
    this_returns = get_returns_in_window(data, 7, 0)
    prev_returns = get_returns_in_window(data, 14, 8)
    return {
        "assignments": {"this": len(this_assigns), "prev": len(prev_assigns)},
        "new_joiner_assigns": {
            "this": sum(1 for a in this_assigns if str(a.get("New Joiner/Replacement", "")).lower() == "new joiner"),
            "prev": sum(1 for a in prev_assigns if str(a.get("New Joiner/Replacement", "")).lower() == "new joiner"),
        },
        "replacements": {"this": len(this_repl), "prev": len(prev_repl)},
        "returns": {"this": len(this_returns), "prev": len(prev_returns)},
    }


# ---------------------------------------------------------------------------
# Snapshot helpers (for stock/week-over-week comparison)
# ---------------------------------------------------------------------------

def get_procurement_runway(data: dict, lookback_weeks: int = 4) -> dict:
    """Estimate weeks of laptop stock at current assignment rate.

    Uses new-joiner assignments over the past `lookback_weeks` weeks as the
    consumption signal (replacements pull from stock too, but new joiners are
    the dominant driver). Returns dict with avg_per_week, stock_ready, weeks.
    """
    cutoff = TODAY - timedelta(days=lookback_weeks * 7)
    count = 0
    for row in data["history"]:
        d = parse_date(row.get("Assigned Date"))
        if d and cutoff <= d <= TODAY:
            atype = str(row.get("New Joiner/Replacement", "")).lower()
            if "joiner" in atype:
                count += 1
    avg_per_week = count / lookback_weeks if lookback_weeks else 0
    stock_ready = len(data["in_stock"])
    weeks: Optional[float] = None
    if avg_per_week > 0:
        weeks = round(stock_ready / avg_per_week, 1)
    return {
        "avg_per_week": round(avg_per_week, 1),
        "stock_ready": stock_ready,
        "weeks": weeks,
    }


def get_spend_pace(data: dict) -> dict:
    """Compare current-month laptop spend (actual) to planned budget.

    Pulls planned total from procurement_plan "Laptop procurement plan" sheet.
    Returns dict with planned, actual, pct_used (or None if planned=0).
    """
    laptop_spend = get_laptop_spend(data)
    actual = laptop_spend["total_spend"]            # USD
    actual_inr = laptop_spend.get("total_spend_inr", 0.0)

    # The plan sheet has individual line items AND summary rows ('Total Estimated
    # Cost', 'Laptop (Exit Employees)', 'Final Estimated Cost'). The authoritative
    # annual budget is the 'Total Estimated Cost' row (the planned figure). We do
    # NOT use 'Final Estimated Cost' — that row now folds in unplanned/ad-hoc
    # spend — nor the raw sum of every Total Price cell (which would multi-count).
    final_cost = total_est = None
    line_sum = 0.0
    for row in data.get("proc_plan", []):
        label = " ".join(" ".join(str(v).split()) for v in row.values()
                          if isinstance(v, str)).lower()
        tp = _to_number(row.get("Total Price (INR)"))
        if "total estimated cost" in label:
            if tp is not None:
                total_est = tp
        elif "final estimated cost" in label:
            if tp is not None:
                final_cost = tp
        elif (tp is not None and "exit" not in label
              and "unplan" not in label and "estimated cost" not in label):
            line_sum += tp
    if total_est is not None:
        planned_inr = total_est
    elif final_cost is not None:
        planned_inr = final_cost
    else:
        planned_inr = line_sum

    planned = inr_to_usd(planned_inr)               # USD, to match `actual`
    # Planned is annual; divide by 12 for a per-month figure.
    monthly_planned = planned / 12 if planned else 0
    monthly_planned_inr = planned_inr / 12 if planned_inr else 0
    pct = None
    if monthly_planned > 0:
        pct = round((actual / monthly_planned) * 100, 0)
    return {
        "actual": actual,
        "actual_inr": actual_inr,
        "monthly_planned": monthly_planned,
        "annual_planned": planned,
        "monthly_planned_inr": monthly_planned_inr,
        "annual_planned_inr": planned_inr,
        "pct_used": pct,
    }


def get_health_check(data: dict, prev_snap: Optional[dict] = None) -> dict:
    """Compute traffic-light status for top dashboard tiles."""
    aging = get_aging_laptops(data)
    critical = sum(1 for a in aging if a["priority"] == "Critical")
    runway = get_procurement_runway(data)
    joiners_7 = get_upcoming_joiners(data, 7)
    stock_ready = len(data["in_stock"])
    pace = get_spend_pace(data)

    # Stock: based on runway weeks
    if runway["weeks"] is None:
        stock_status = "🟢"  # no recent consumption
    elif runway["weeks"] >= 4:
        stock_status = "🟢"
    elif runway["weeks"] >= 2:
        stock_status = "🟡"
    else:
        stock_status = "🔴"

    # Aging: based on critical count
    if critical == 0:
        aging_status = "🟢"
    elif critical <= 5:
        aging_status = "🟡"
    else:
        aging_status = "🔴"

    # Joiner prep: stock vs joiners next 7 days
    if not joiners_7:
        joiner_status = "🟢"
    elif stock_ready >= len(joiners_7):
        joiner_status = "🟢"
    elif stock_ready >= len(joiners_7) - 2:
        joiner_status = "🟡"
    else:
        joiner_status = "🔴"

    # Spend: vs monthly budget pace
    if pace["pct_used"] is None:
        spend_status = "🟢"
    elif pace["pct_used"] < 90:
        spend_status = "🟢"
    elif pace["pct_used"] < 110:
        spend_status = "🟡"
    else:
        spend_status = "🔴"

    return {
        "stock": stock_status,
        "aging": aging_status,
        "joiner_prep": joiner_status,
        "spend": spend_status,
        "_critical_aging": critical,
        "_runway_weeks": runway["weeks"],
        "_joiners_7": len(joiners_7),
        "_stock_ready": stock_ready,
        "_pace_pct": pace["pct_used"],
    }


def get_risk_callouts(data: dict, hc: dict) -> list[str]:
    """Auto-detect risks worth highlighting at the top of the report."""
    risks: list[str] = []
    runway = hc["_runway_weeks"]
    joiners_7 = hc["_joiners_7"]
    stock_ready = hc["_stock_ready"]
    critical = hc["_critical_aging"]
    pace_pct = hc["_pace_pct"]

    # Joiner stock shortage
    if joiners_7 > stock_ready:
        gap = joiners_7 - stock_ready
        risks.append(f"🔴 *{joiners_7} joiners next week*, only {stock_ready} laptops in stock — order *{gap}* laptops")
    elif joiners_7 and stock_ready - joiners_7 <= 2:
        risks.append(f"🟡 Tight stock: {stock_ready} laptops vs {joiners_7} joiners next week")

    # Stock runway
    if runway is not None and runway < 2:
        risks.append(f"🔴 *Stock runway: {runway} weeks* at current assignment pace")
    elif runway is not None and runway < 4:
        risks.append(f"🟡 Stock runway: {runway} weeks — plan procurement soon")

    # Critical aging
    if critical > 5:
        risks.append(f"🟡 *{critical} laptops >4 yrs* still assigned — schedule replacements")

    # Spend pace
    if pace_pct is not None and pace_pct > 110:
        risks.append(f"🔴 Laptop spend pace: *{pace_pct:.0f}%* of monthly budget — over by {pace_pct - 100:.0f}%")

    return risks


def current_snapshot(data: dict) -> dict:
    """Build a snapshot of key metrics for persistence."""
    total_assigned = sum(1 for r in data["assigned"] if is_truly_assigned(r))
    aging = get_aging_laptops(data)
    laptop_spend = get_laptop_spend(data)
    total_app, _, _, _ = get_current_month_spend(data)
    joiners_30 = get_upcoming_joiners(data, 30)
    return {
        "date": TODAY.isoformat(),
        "stock_ready": len(data["in_stock"]),
        "stock_backup": len(data["backup"]),
        "total_assigned": total_assigned,
        "total_laptops": total_assigned + len(data["in_stock"]) + len(data["backup"]),
        "aging_count": len(aging),
        "laptop_spend_month": laptop_spend["total_spend"],
        "laptop_joiners_month": laptop_spend["total_joiners"],
        "app_spend_month": total_app,
        "joiners_next_30": len(joiners_30),
    }


def load_previous_snapshot() -> Optional[dict]:
    """Read the snapshot saved by the previous run, if any."""
    path = SNAPSHOT_DIR / "latest.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def save_snapshot(snap: dict) -> None:
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    (SNAPSHOT_DIR / "latest.json").write_text(
        json.dumps(snap, indent=2), encoding="utf-8"
    )


def _delta_icon(cur: float, prev: float, good_is_up: bool = False) -> str:
    if cur == prev:
        return "➖"
    up = cur > prev
    if good_is_up:
        return "🟢 ↑" if up else "🔴 ↓"
    return "🔴 ↑" if up else "🟢 ↓"


def _fmt_delta(cur: float, prev: Optional[float], as_int: bool = True) -> str:
    if prev is None:
        return "—"
    diff = cur - prev
    if diff == 0:
        return "±0"
    sign = "+" if diff > 0 else ""
    if as_int:
        return f"{sign}{int(diff)}"
    return f"{sign}{diff:,.0f}"


def _aging_remark(priority: str, warranty_end: Optional[dt.date]) -> str:
    """Auto-generate an actionable remark for an aging laptop based on its
    replacement priority and warranty status. No per-employee notes exist in
    the source data, so these are derived deterministically."""
    parts = []
    if priority == "Critical":
        parts.append("Replace now (>4 yrs)")
    else:  # High
        parts.append("Replace within 30 days")
    if warranty_end:
        days = (warranty_end - TODAY).days
        if days < 0:
            parts.append(f"Warranty expired {warranty_end.strftime('%b %Y')}")
        elif days <= 90:
            parts.append(f"Warranty ends {warranty_end.strftime('%b %Y')}")
        else:
            parts.append("In warranty")
    else:
        parts.append("Warranty date missing")
    return " · ".join(parts)


def get_aging_laptops(data: dict) -> list[dict]:
    """Return assigned laptops older than 3.5 years, sorted oldest first."""
    aging = []
    for row in data["assigned"]:
        if not is_truly_assigned(row):
            continue
        dt = parse_date(row.get("Warranty Start Date"))
        if dt and (TODAY - dt).days > AGE_THRESHOLD_DAYS:
            priority = "Critical" if age_years(dt) > 4 else "High"
            warranty_end = parse_date(row.get("Warranty End Date"))
            aging.append({
                "employee": row.get("Employee Name", "Unknown"),
                "department": row.get("Department", ""),
                "tag": row.get("Laptop Asset Tag", ""),
                "make": row.get("Laptop Make", ""),
                "model": row.get("Laptop Model", ""),
                "start_date": dt,
                "warranty_end": warranty_end,
                "age_years": round(age_years(dt), 1),
                "priority": priority,
                "remark": _aging_remark(priority, warranty_end),
            })
    aging.sort(key=lambda x: x["age_years"], reverse=True)
    return aging


def build_employee_directory(data: dict) -> str:
    """A full roster of who currently holds which laptop — so the bot can answer
    per-person questions ('what laptop does X have?', 'is X's laptop old?')."""
    rows = []
    for row in data["assigned"]:
        if not is_truly_assigned(row):
            continue
        emp = str(row.get("Employee Name", "") or "").strip()
        if not emp:
            continue
        start = parse_date(row.get("Warranty Start Date"))
        wend = parse_date(row.get("Warranty End Date"))
        make = str(row.get("Laptop Make", "") or "").strip()
        model = str(row.get("Laptop Model", "") or "").strip()
        serial = str(row.get("Laptop Serial Number", "") or "").strip()
        tag = str(row.get("Laptop Asset Tag", "") or "").strip()
        age = round(age_years(start), 1) if start else None
        due = bool(start and (TODAY - start).days > AGE_THRESHOLD_DAYS)
        rows.append({
            "emp": emp, "dept": str(row.get("Department", "") or "").strip(),
            "laptop": f"{make} {model}".strip(), "id": serial or tag,
            "age": age, "wend": wend, "due": due,
        })
    rows.sort(key=lambda r: r["emp"].lower())
    L = [f"# EMPLOYEE LAPTOP DIRECTORY ({len(rows)} laptops currently assigned)",
         "Who holds which laptop. 'Replace due?' = older than 3.5 years.\n",
         "| Employee | Department | Laptop | Serial/Tag | Age (yrs) | Warranty ends | Replace due? |",
         "|---|---|---|---|---|---|---|"]
    for r in rows:
        L.append(f"| {r['emp']} | {r['dept'] or '—'} | {r['laptop'] or '—'} | {r['id'] or '—'} "
                 f"| {r['age'] if r['age'] is not None else '—'} "
                 f"| {r['wend'].strftime('%b %Y') if r['wend'] else '—'} "
                 f"| {'YES' if r['due'] else 'no'} |")
    return "\n".join(L)


def _bot_joiners_section(data: dict) -> str:
    """Upcoming joiners (next 30 days) with laptop needed + onboarding checklist."""
    joiners = get_joiners_with_laptop_needs(data, 30)

    def _norm(s):
        return str(s or "").strip().lower()

    by_name = {}
    for row in data.get("checklist", []):
        name = row.get("Name ", row.get("Name", ""))
        if name:
            by_name[_norm(name)] = row

    L = [f"# UPCOMING JOINERS — next 30 days ({len(joiners)})",
         "Joining date, role, laptop needed, and onboarding-checklist status per joiner. "
         "Compare laptop needs against spare stock to judge readiness.\n"]
    if not joiners:
        L.append("None joining in the next 30 days.")
        return "\n".join(L)
    for j in joiners:
        row = by_name.get(_norm(j["name"]))
        if not row:
            status = "onboarding checklist not found"
        else:
            done, pending = [], []
            for col in ONBOARDING_CHECKLIST_COLS:
                v = row.get(col)
                if v is None:
                    continue
                (done if str(v).strip().lower() not in ("", "none", "no", "pending", "0")
                 else pending).append(col)
            total = len(done) + len(pending)
            status = (f"{len(done)}/{total} checklist done"
                      + (f"; pending: {', '.join(pending)}" if pending else "; all done"))
        fast = get_fastest_delivery_for_department(data, j["department"])
        if fast and fast["fastest"]:
            f = fast["fastest"]
            if f["days"] is not None and f["days"] <= j["days_until"]:
                deliv = (f" · laptop can arrive in time if ordered now "
                         f"(fastest {f['vendor']} ~{f['text']})")
            else:
                deliv = (f" · ⚠️ delivery risk — fastest ~{f['text']} ({f['vendor']}) "
                         f"vs joins in {j['days_until']}d")
        else:
            deliv = ""
        L.append(f"- *{j['name']}* — {j['department'] or '—'}, {j['designation'] or '—'} · "
                 f"DOJ {j['doj'].strftime('%d %b %Y')} (in {j['days_until']}d) · "
                 f"laptop needed: {j['laptop_config'] or 'standard'} · onboarding: {status}{deliv}")
    return "\n".join(L)


def _bot_delivery_section(data: dict) -> str:
    """Combined vendor delivery + payment-terms table — the single table to show
    for 'laptop delivery timelines from vendors'."""
    matrix = get_vendor_delivery_matrix(data)
    if not matrix["rows"]:
        return "# LAPTOP DELIVERY & PAYMENT TERMS\nNo delivery-timeline table found."
    cols = " | ".join(matrix["devices"])
    L = ["# LAPTOP DELIVERY & PAYMENT TERMS (by vendor)",
         "When asked about delivery timelines/vendors, show THIS as one table "
         "(vendors fastest-first), with both the per-device lead time and payment terms.\n",
         f"| Vendor | {cols} | Payment terms |",
         "|" + "---|" * (len(matrix["devices"]) + 2)]
    for r in matrix["rows"]:
        L.append("| " + " | ".join([r["vendor"], *r["cells"], r["terms"]]) + " |")
    return "\n".join(L)


def get_pending_returns(data: dict) -> list[dict]:
    """Laptops still to be returned by employees, from the 'Laptop yet to Return'
    sheet."""
    out = []
    for row in data.get("pending_returns", []):
        emp = str(row.get("Employee Name") or row.get("Username") or "").strip()
        if not emp and not str(row.get("Laptop Serial Number") or "").strip():
            continue
        make = str(row.get("Laptop Make", "") or "").strip()
        model = str(row.get("Laptop Model", "") or "").strip()
        out.append({
            "employee": emp,
            "department": str(row.get("Department", "") or "").strip(),
            "laptop": f"{make} {model}".strip(),
            "serial": str(row.get("Laptop Serial Number") or row.get("Laptop Asset Tag") or "").strip(),
            "remarks": str(row.get("Remarks", "") or "").strip(),
        })
    return out


def _bot_returns_section(data: dict) -> str:
    """Laptop returns + still-pending returns (offboarding)."""
    # Returned laptops. Real columns: Username, Laptop Make/Model, Serial Number,
    # Laptop Tag, Returned Date, Resigned/Replacement.
    items = []
    for row in data.get("returned", []):
        emp = str(row.get("Username") or row.get("Employee Name") or "").strip()
        make = str(row.get("Laptop Make", "") or "").strip()
        model = str(row.get("Laptop Model", "") or "").strip()
        sn = str(row.get("Serial Number") or row.get("Laptop Tag") or "").strip()
        reason = str(row.get("Resigned/Replacement", "") or "").strip()
        items.append((emp, f"{make} {model}".strip(),
                      parse_date(row.get("Returned Date")), sn, reason))
    items.sort(key=lambda x: x[2] or dt.date.min, reverse=True)

    pending = get_pending_returns(data)
    L = [f"# LAPTOPS YET TO RETURN ({len(pending)} pending)",
         "Laptops still to be collected from employees (e.g. exits). "
         "Use for 'how many laptops are yet to be returned?'.\n"]
    if pending:
        L.append("| Employee | Dept | Laptop | Serial/Tag | Remarks |")
        L.append("|---|---|---|---|---|")
        for p in pending:
            L.append(f"| {p['employee'] or '—'} | {p['department'] or '—'} | {p['laptop'] or '—'} "
                     f"| {p['serial'] or '—'} | {p['remarks'] or '—'} |")
    else:
        L.append("None pending.")

    L.append(f"\n# LAPTOP RETURNS — completed ({len(items)} on record, most recent first)")
    L.append("| Employee | Laptop | Returned on | Serial/Tag | Reason |")
    L.append("|---|---|---|---|---|")
    for emp, laptop, d, sn, reason in items[:60]:
        L.append(f"| {emp or '—'} | {laptop or '—'} | {d.strftime('%d %b %Y') if d else '—'} "
                 f"| {sn or '—'} | {reason or '—'} |")
    return "\n".join(L)


_BOT_PERIPHERALS = [("mouse", "Mouse"), ("headset", "Headset"), ("keyboard", "Keyboard"),
                    ("charger", "Charger"), ("docking", "Docking station"), ("monitor", "Monitor")]


def _row_employee(row: dict) -> str:
    for k, v in row.items():
        kl = str(k).strip().lower()
        if (("employee" in kl or kl.startswith("name") or "user" in kl or "assigned to" in kl)
                and str(v or "").strip()):
            return str(v).strip()
    return ""


def _row_detail(row: dict) -> str:
    parts = []
    for k, v in row.items():
        kl = str(k).strip().lower()
        if any(t in kl for t in ("make", "model", "brand", "type", "size")) and str(v or "").strip():
            parts.append(str(v).strip())
    return " ".join(parts[:2])


def _bot_peripherals_section(data: dict) -> str:
    """Per-person accessories (monitor, headset, keyboard, mouse, docking, charger)."""
    by_person = defaultdict(list)
    for key, label in _BOT_PERIPHERALS:
        for row in data.get(key, []):
            emp = _row_employee(row)
            if not emp:
                continue
            detail = _row_detail(row)
            by_person[emp].append(f"{label}{f' ({detail})' if detail else ''}")
    L = [f"# PERIPHERALS BY PERSON ({len(by_person)} people with accessories on record)",
         "Who holds which accessories. Use for 'does X have a monitor/headset?'.\n",
         "| Employee | Accessories |", "|---|---|"]
    for emp in sorted(by_person, key=str.lower):
        L.append(f"| {emp} | {', '.join(by_person[emp])} |")
    return "\n".join(L)


def _bot_tickets_section(data: dict) -> str:
    """Open IT tickets with requester + days open (for 'any tickets for X?')."""
    issues = get_it_issues(data)
    open_issues = [i for i in issues["issues"] if i["is_open"]]
    L = [f"# IT TICKETS — open ({len(open_issues)})",
         "Use for 'any IT requests for <person>?' and 'how long has it been pending?'.\n",
         "| Issue | Raised by | Owner | Status | Days open | Why pending |",
         "|---|---|---|---|---|---|"]
    for i in open_issues:
        days = (TODAY - i["date"]).days if i["date"] else None
        L.append(f"| {i['issue']} | {i['raised_by'] or '—'} | {i['owner'] or '—'} "
                 f"| {i['status'] or 'Open'} | {days if days is not None else '—'} | {i['remark']} |")
    return "\n".join(L)


def _bot_software_section(data: dict) -> str:
    """Software/licence inventory + upcoming renewals + budget vs actual."""
    inv = get_software_inventory(data)
    renewals = get_upcoming_renewals(data, 60)
    bva = get_budget_vs_actual(data)
    L = [f"# SOFTWARE & LICENSES ({len(inv)} subscriptions)",
         "Apps/subscriptions, monthly cost, owner, and renewal date.\n",
         "| Application | Cost (this month) | Frequency | Owner/Dept | Renews |",
         "|---|---|---|---|---|"]
    for i in inv:
        L.append(f"| {i['app']} | {fmt_usd(i['cost']) if i['cost'] is not None else '—'} "
                 f"| {i['frequency'] or '—'} | {i['dept'] or '—'} "
                 f"| {i['renewal'].strftime('%d %b %Y') if i['renewal'] else '—'} |")
    L.append(f"\n## RENEWALS DUE — next 60 days ({len(renewals)})")
    for r in renewals:
        L.append(f"- {r['renewal'].strftime('%d %b %Y')}: {r['app']} "
                 f"({fmt_usd(r['cost']) if r['cost'] is not None else '—'})")
    if not renewals:
        L.append("- None.")
    L.append("\n## LAPTOP PROCUREMENT: PLAN vs ACTUAL (amounts in INR)")
    L.append(f"- Planned laptop procurement for the year (FY26, from the 'Laptop procurement "
             f"plan' sheet): {fmt_inr_full(bva['laptop_annual_inr'])} — about "
             f"{fmt_inr_full(bva['laptop_monthly_inr'])}/month. NOTE: the per-month figure is "
             f"the planned annual total ÷ 12, NOT a separately-approved budget.")
    pct = (f" — {bva['laptop_pct_of_monthly']:.0f}% of that planned monthly amount"
           if bva['laptop_pct_of_monthly'] is not None else "")
    models = ", ".join(f"{m['model']} ×{m['units']}" for m in bva["laptop_models"])
    L.append(f"- Laptops procured this month: {bva['laptops_this_month']}"
             + (f" — {models}" if models else ""))
    L.append(f"- Laptop spend this month: {fmt_inr_full(bva['laptop_spend_inr'])}{pct}")
    if bva["purchases_this_month"]:
        L.append("- Laptops in the purchase register this month (model · purchase date · serial):")
        for p in bva["purchases_this_month"]:
            L.append(f"  • {(p['brand'] + ' ' + p['model']).strip() or '—'} · "
                     f"{p['date'].strftime('%d %b %Y') if p['date'] else 'date not recorded'} · "
                     f"{p['serial'] or '—'}")
    # Month-by-month laptop spend (FY26 to date) so YTD / 'Jan–Jun' questions work.
    hist = bva["spend_history"]
    if hist["months"]:
        L.append("- Laptop spend by month so far this year (FY26), ₹:")
        for m in hist["months"]:
            mdl = ", ".join(f"{x['model']} ×{x['units']}" for x in m.get("models", []))
            L.append(f"  • {m['month']}: {m['units']} laptop(s), {fmt_inr_full(m['spend_inr'])}"
                     + (f" ({mdl})" if mdl else ""))
        L.append(f"  • YTD total ({hist['months'][0]['month']}–{hist['months'][-1]['month']}): "
                 f"{hist['ytd_units']} laptop(s), {fmt_inr_full(hist['ytd_spend_inr'])}")
        elapsed = len(hist["months"])
        planned_td = bva["laptop_monthly_inr"] * elapsed
        L.append(f"  • Planned for these {elapsed} months (₹monthly × {elapsed}): "
                 f"{fmt_inr_full(planned_td)}; actual {fmt_inr_full(hist['ytd_spend_inr'])}.")
        diff = planned_td - hist["ytd_spend_inr"]
        if diff > 0:
            L.append(f"  • SAVED so far: {fmt_inr_full(diff)} under the planned budget "
                     f"(spent {fmt_inr_full(hist['ytd_spend_inr'])} of a planned "
                     f"{fmt_inr_full(planned_td)}). When spend is below budget, this is "
                     f"the saving — state it as money saved.")
        elif diff < 0:
            L.append(f"  • OVER budget so far by {fmt_inr_full(-diff)} "
                     f"(spent {fmt_inr_full(hist['ytd_spend_inr'])} vs a planned "
                     f"{fmt_inr_full(planned_td)}).")
        else:
            L.append("  • Exactly on the planned budget so far (no saving or overspend).")
    # Which vendor each laptop was bought from — from the purchase register's
    # 'Purchased From' column, so "from which vendor did we buy?" is answerable.
    bv = get_purchases_by_vendor(data)
    if bv["vendors"]:
        rng = ""
        if bv["date_from"] and bv["date_to"]:
            rng = (f" ({bv['date_from'].strftime('%b %Y')}–"
                   f"{bv['date_to'].strftime('%b %Y')})")
        L.append(f"- Laptops by vendor (who we bought from), from the 'New Laptops "
                 f"purchased' register{rng} — {bv['total']} laptop(s) total:")
        for v in bv["vendors"]:
            mdl = ", ".join(f"{m['model']} ×{m['units']}" for m in v["models"])
            L.append(f"  • {v['vendor']}: {v['count']} laptop(s)"
                     + (f" — {mdl}" if mdl else ""))
        L.append("  NOTE: this register is the source for which vendor supplied each "
                 "laptop; its laptop count may differ from the monthly spend tracker "
                 "above because the two sheets are maintained separately.")
    L.append(f"- Software & licenses this month: {fmt_usd(bva['software_this_month'])}")
    shist = get_software_spend_history(data)
    if shist["months"]:
        L.append("- Software/subscription spend by month so far this year ($):")
        for m in shist["months"]:
            L.append(f"  • {m['month']}: {fmt_usd(m['spend'])}")
        L.append(f"  • YTD total: {fmt_usd(shist['ytd'])}")
    return "\n".join(L)


def _bot_joiners_history_section(data: dict) -> str:
    """New joiners + laptop assignments (new-joiner vs replacement) per month."""
    h = get_joiners_history(data)
    a = get_assignment_history(data)
    by_m = {x["month"]: x for x in a["months"]}
    L = [f"# NEW JOINERS & LAPTOP REPLACEMENTS THIS YEAR",
         f"People joined (from the joiners sheet) and laptops assigned each month, "
         f"split into new-joiner vs replacement (from the asset history). "
         f"YTD: {h['ytd']} joined, {a['ytd_new_joiner']} new-joiner laptops, "
         f"{a['ytd_replacement']} replacement laptops.\n",
         "| Month | People joined | New-joiner laptops | Replacements |",
         "|---|---|---|---|"]
    for m in h["months"]:
        am = by_m.get(m["month"], {"new_joiner": 0, "replacement": 0})
        L.append(f"| {m['month']} | {m['count']} | {am['new_joiner']} | {am['replacement']} |")
    L.append(f"| **YTD** | **{h['ytd']}** | **{a['ytd_new_joiner']}** | **{a['ytd_replacement']}** |")
    # Names per month, so "who joined in May?" can be answered, not just a count.
    named = [m for m in h["months"] if m["names"]]
    if named:
        L.append("\nWho joined each month (names, for 'who joined in <month>?'):")
        for m in named:
            L.append(f"- {m['month']}: {', '.join(m['names'])}")
    return "\n".join(L)


def _bot_unplanned_section(data: dict) -> str:
    """Unplanned / ad-hoc spends — a separate section, from a dedicated tab OR
    the 'UnPlaned' rows in the procurement plan (kept apart from the budget)."""
    u = get_unplanned_spends(data)
    mode = u.get("mode", "none")
    if mode == "none":
        return ("# UNPLANNED / AD-HOC SPENDS\n"
                "No unplanned spends are recorded right now (no 'Unplanned' tab, "
                "and no 'UnPlaned' rows in the procurement plan).")

    # Rows recorded inside the Laptop procurement plan sheet.
    if mode == "plan_rows":
        items = u["items"]
        L = [f"# UNPLANNED / AD-HOC SPENDS ({len(items)} item(s))",
             "Unplanned / ad-hoc laptop replacements recorded in the 'Laptop "
             "procurement plan' sheet (the 'UnPlaned Replacements' rows) — kept "
             "SEPARATE from the planned Total Estimated Cost budget.\n",
             "| Category | Model | Qty | Amount (INR) |",
             "|---|---|---|---|"]
        for it in items:
            qty = "—" if it["quantity"] is None else str(it["quantity"])
            amt = (fmt_inr_full(it["amount_inr"])
                   if it["amount_inr"] is not None else "—")
            L.append(f"| {it['category'] or '—'} | {it['model'] or '—'} "
                     f"| {qty} | {amt} |")
        if u["total"] is not None:
            L.append(f"\nTotal unplanned spend: {fmt_inr_full(u['total'])}.")
        else:
            L.append("\n(Amounts for these unplanned items aren't filled in on the "
                     "sheet — only the models/quantities are recorded.)")
        return "\n".join(L)

    # Dedicated tab — reproduced verbatim.
    headers, amount_col = u["headers"], u["amount_col"]

    def _cell(h, v):
        if v is None or str(v).strip() == "":
            return "—"
        if h == amount_col:
            n = _to_number(v)
            return fmt_inr_full(n) if n is not None else " ".join(str(v).split())
        if "date" in str(h).lower():
            d = parse_date(v)
            if d:
                return d.strftime("%d %b %Y")
        if isinstance(v, (dt.datetime, dt.date)):
            return v.strftime("%d %b %Y")
        return " ".join(str(v).split()).replace("|", "/")

    L = [f"# UNPLANNED / AD-HOC SPENDS ({len(u['rows'])} item(s))",
         f"Spends recorded OUTSIDE the planned budget, from the '{u['sheet']}' "
         f"sheet — shown exactly as recorded and kept SEPARATE from the planned "
         f"laptop procurement budget.\n",
         "| " + " | ".join(headers) + " |",
         "|" + "---|" * len(headers)]
    for r in u["rows"]:
        L.append("| " + " | ".join(_cell(h, r.get(h)) for h in headers) + " |")
    if u["total"] is not None:
        L.append(f"\nTotal unplanned spend (sum of '{amount_col}'): "
                 f"{fmt_inr_full(u['total'])}.")
    return "\n".join(L)


def _rows_to_table(rows: list[dict], max_rows: int = 80) -> list[str]:
    """Render arbitrary sheet rows verbatim as a markdown table. Keeps only
    columns that have a real header OR carry data (blank-header columns become
    'Item'), drops fully-empty rows, and caps the row count. Used to surface
    tabs whose exact columns we don't want to reinterpret."""
    rows = [r for r in rows
            if any(str(v).strip() for v in r.values() if v is not None)]
    if not rows:
        return []
    cols: list = []
    for r in rows:
        for k in r.keys():
            if k not in cols:
                cols.append(k)
    keep = [c for c in cols
            if not str(c).startswith("col_")
            or any(str(r.get(c) or "").strip() for r in rows)]

    def _cell(v):
        if v is None or str(v).strip() == "":
            return "—"
        if isinstance(v, (dt.datetime, dt.date)):
            return v.strftime("%d %b %Y")
        return " ".join(str(v).split()).replace("|", "/")

    hdr = ["Item" if str(c).startswith("col_") else str(c) for c in keep]
    lines = ["| " + " | ".join(hdr) + " |", "|" + "---|" * len(keep)]
    for r in rows[:max_rows]:
        lines.append("| " + " | ".join(_cell(r.get(c)) for c in keep) + " |")
    if len(rows) > max_rows:
        lines.append(f"\n(+{len(rows) - max_rows} more row(s) not shown.)")
    return lines


def _bot_sold_section(data: dict) -> str:
    """Laptops sold / disposed (retired from the company), reproduced verbatim
    from the 'Laptops sold' tab. Answers 'have we sold any laptops?'."""
    rows = data.get("sold", [])
    body = _rows_to_table(rows)
    L = [f"# LAPTOPS SOLD / DISPOSED ({len([1 for r in rows if any(str(v).strip() for v in r.values() if v is not None)])})",
         "Laptops that have been sold or disposed of (retired from use). Use for "
         "'have we sold any laptops?', 'which laptops were disposed/retired?'. "
         "These are company-owned laptops we let go of — NOT sales to customers."]
    if not body:
        L.append("\nNo laptops are recorded as sold or disposed in the current data.")
        return "\n".join(L)
    L.append("")
    L.extend(body)
    return "\n".join(L)


def _bot_other_stock_section(data: dict) -> str:
    """Non-laptop spare assets in stock (USB hubs, adapters, etc.) with their
    quantities, from the 'Other Assets Instock' tab."""
    rows = data.get("other_stock", [])
    body = _rows_to_table(rows)
    L = ["# OTHER ASSETS IN STOCK",
         "Spare non-laptop assets held in stock (e.g. USB hubs, adapters, cables) "
         "with quantities. Use for 'how many spare <item> do we have?'."]
    if not body:
        L.append("\nNo other-asset stock is recorded.")
        return "\n".join(L)
    L.append("")
    L.extend(body)
    return "\n".join(L)


def _bot_backup_section(data: dict) -> str:
    """Backup / spare laptops from the '3 years old' pool — older standby
    machines, surfaced with make/model/config so 'which backup laptops do we
    have?' is answerable (the main ready stock is in the report's stock section)."""
    rows = data.get("backup", [])
    body = _rows_to_table(rows)
    n = len([1 for r in rows
             if any(str(v).strip() for v in r.values() if v is not None)])
    L = [f"# BACKUP / SPARE LAPTOPS — 3+ years old ({n})",
         "Older spare laptops kept as standby/backup, separate from the main ready "
         "stock. Use for 'how many backup/old spare laptops do we have?' and "
         "'which models are in the backup pool?'."]
    if not body:
        L.append("\nNo backup laptops are recorded.")
        return "\n".join(L)
    L.append("")
    L.extend(body)
    return "\n".join(L)


def _bot_config_section(data: dict) -> str:
    """Standard laptop configuration per department/role (the spec a new joiner
    is given), so 'what laptop/spec does a <role> get?' is answerable."""
    rows = data.get("configuration", [])
    body = _rows_to_table(rows)
    L = ["# STANDARD LAPTOP CONFIGURATIONS (by department / role)",
         "The standard laptop spec each department/role is given (device type, RAM, "
         "processor). Use for 'what laptop/spec does a <role/department> get?'."]
    if not body:
        L.append("\nNo standard-configuration table is recorded.")
        return "\n".join(L)
    L.append("")
    L.extend(body)
    return "\n".join(L)


def _norm_serial(s) -> str:
    """Normalise a serial number for matching (trim + lowercase)."""
    return str(s or "").strip().lower()


def _history_by_serial(data: dict) -> dict:
    """Index the Asset History tab by normalised serial number, keeping the most
    recent record per serial (latest Assigned Date wins). Asset History is the
    only tab carrying the 'New Joiner/Replacement' flag. Shared by the
    assignment-type and vendor-purchase joins."""
    out: dict = {}
    for h in data.get("history", []):
        sn = _norm_serial(h.get("Serial Number") or h.get("Laptop Tag"))
        if not sn:
            continue
        d = parse_date(h.get("Assigned Date"))
        rec = {"type": str(h.get("New Joiner/Replacement", "") or "").strip(),
               "date": d,
               "user": str(h.get("Username", "") or "").strip(),
               "make": str(h.get("Laptop Make", "") or "").strip(),
               "model": str(h.get("Laptop Model", "") or "").strip(),
               "serial": str(h.get("Serial Number") or h.get("Laptop Tag") or "").strip()}
        prev = out.get(sn)
        if prev is None or (d and (prev["date"] is None or d >= prev["date"])):
            out[sn] = rec
    return out


def _assigned_by_serial(data: dict) -> dict:
    """Index currently-assigned laptops by normalised serial → {employee, dept}."""
    out: dict = {}
    for row in data.get("assigned", []):
        if not is_truly_assigned(row):
            continue
        sn = _norm_serial(row.get("Laptop Serial Number"))
        if sn:
            out[sn] = {"employee": str(row.get("Employee Name", "") or "").strip(),
                       "dept": str(row.get("Department", "") or "").strip()}
    return out


def get_assignment_types(data: dict) -> dict:
    """Join currently-assigned laptops to the Asset History tab BY SERIAL NUMBER
    to label each as a new-joiner or replacement assignment. Returns:
      matched      -> assigned laptops with a history record (incl. type + date)
      unmatched    -> assigned laptops with no history record (type unknown)
      history_only -> history records whose serial isn't currently assigned
    """
    def _ns(s) -> str:
        return _norm_serial(s)

    hist_by_serial = _history_by_serial(data)
    matched, unmatched, used = [], [], set()
    for row in data.get("assigned", []):
        if not is_truly_assigned(row):
            continue
        emp = str(row.get("Employee Name", "") or "").strip()
        if not emp:
            continue
        sn = _ns(row.get("Laptop Serial Number"))
        item = {"employee": emp,
                "dept": str(row.get("Department", "") or "").strip(),
                "make": str(row.get("Laptop Make", "") or "").strip(),
                "model": str(row.get("Laptop Model", "") or "").strip(),
                "serial": str(row.get("Laptop Serial Number", "") or "").strip(),
                "tag": str(row.get("Laptop Asset Tag", "") or "").strip()}
        h = hist_by_serial.get(sn) if sn else None
        if h:
            used.add(sn)
            item["type"] = h["type"] or None
            item["date"] = h["date"]
            matched.append(item)
        else:
            item["type"], item["date"] = None, None
            unmatched.append(item)

    history_only = [r for sn, r in hist_by_serial.items() if sn not in used]
    history_only.sort(key=lambda r: r["date"] or dt.date(1900, 1, 1), reverse=True)
    return {"matched": matched, "unmatched": unmatched, "history_only": history_only}


def _bot_assignment_type_section(data: dict) -> str:
    """Per-laptop new-joiner vs replacement, matched by serial number between the
    assigned-laptop sheet and the Asset History tab."""
    a = get_assignment_types(data)
    matched, unmatched, hist_only = a["matched"], a["unmatched"], a["history_only"]
    L = ["# LAPTOP ASSIGNMENTS BY SERIAL — new joiner vs replacement",
         "Each currently-assigned laptop matched to the Asset History tab BY SERIAL "
         "NUMBER to show whether it was given as a NEW-JOINER setup or a REPLACEMENT. "
         "Use for 'which laptops were replacements vs new joiners?' and 'was <name>'s "
         "laptop a replacement?'. The new-joiner/replacement flag exists ONLY in Asset "
         "History; an assigned laptop with no matching serial there shows 'not recorded'."]

    def _laptop(d):
        return (f"{d['make']} {d['model']}").strip() or "—"

    if matched:
        L.append(f"\n## Assigned laptops matched to history ({len(matched)})")
        L.append("| Employee | Laptop | Serial | New joiner / Replacement | Assigned |")
        L.append("|---|---|---|---|---|")
        for m in matched:
            date = m["date"].strftime("%d %b %Y") if m["date"] else "—"
            L.append(f"| {m['employee']} | {_laptop(m)} | {m['serial'] or '—'} "
                     f"| {m['type'] or 'not recorded'} | {date} |")
    if unmatched:
        L.append(f"\n## Assigned laptops with NO Asset-History record ({len(unmatched)})")
        L.append("New-joiner/replacement is not recorded for these (no matching serial "
                 "in Asset History).")
        L.append("| Employee | Laptop | Serial |")
        L.append("|---|---|---|")
        for u in unmatched:
            L.append(f"| {u['employee']} | {_laptop(u)} | {u['serial'] or '—'} |")
    if hist_only:
        L.append(f"\n## Asset-History records not tied to a current assignment ({len(hist_only)})")
        L.append("Past or other assignments from Asset History — the laptop isn't in "
                 "the current assigned list (e.g. since returned or reassigned).")
        L.append("| Person | Laptop | Serial | New joiner / Replacement | Assigned |")
        L.append("|---|---|---|---|---|")
        for h in hist_only:
            date = h["date"].strftime("%d %b %Y") if h["date"] else "—"
            L.append(f"| {h['user'] or '—'} | {_laptop(h)} | {h['serial'] or '—'} "
                     f"| {h['type'] or 'not recorded'} | {date} |")
    if not (matched or unmatched or hist_only):
        L.append("\nNo assigned laptops or history records are available.")
    return "\n".join(L)


def get_vendor_purchase_details(data: dict) -> dict:
    """Every laptop in the 'New Laptops purchased' register, grouped by the vendor
    it was bought from, and joined BY SERIAL NUMBER to who currently holds it
    (assigned sheet) and whether that was a new-joiner or replacement assignment
    (Asset History). Lets the bot answer 'how many laptops from <vendor>, with
    serials, who has them, and is it a replacement or new joiner?'.
    Returns {vendors:[{vendor,count,laptops:[...]}], total}."""
    assigned = _assigned_by_serial(data)
    hist = _history_by_serial(data)
    by_vendor: dict[str, list] = {}
    total = 0
    for row in data.get("purchased", []):
        brand = str(row.get("Brand", "") or "").strip()
        model = str(row.get("Model", "") or "").strip()
        serial = str(row.get("Serial no", "") or "").strip()
        if not (brand or model or serial):
            continue  # skip blank/spacer rows
        vendor = str(row.get("Purchased From", "") or "").strip() or "Not recorded"
        sn = _norm_serial(serial)
        a = assigned.get(sn) if sn else None
        h = hist.get(sn) if sn else None
        by_vendor.setdefault(vendor, []).append({
            "serial": serial or "—",
            "laptop": (f"{brand} {model}").strip() or "—",
            "assignee": a["employee"] if a else None,
            "dept": a["dept"] if a else None,
            "type": (h["type"] if h and h["type"] else None),
            "date": parse_date(row.get("Warranty Start Date")),
        })
        total += 1
    vendors = [{"vendor": k, "count": len(v), "laptops": v}
               for k, v in by_vendor.items()]
    vendors.sort(key=lambda x: -x["count"])
    return {"vendors": vendors, "total": total}


def _bot_vendor_purchase_section(data: dict) -> str:
    """Per-vendor purchased laptops with serial, current assignee and new-joiner/
    replacement — the detailed answer to 'how many laptops from <vendor>?'."""
    vp = get_vendor_purchase_details(data)
    L = [f"# LAPTOP PURCHASES BY VENDOR — serial · assignee · type ({vp['total']})",
         "Every laptop in the 'New Laptops purchased' register, grouped by the vendor "
         "it was bought from, each joined BY SERIAL NUMBER to who holds it now and "
         "whether that was a new-joiner or replacement assignment. Use for 'how many "
         "laptops did we buy from <vendor>, with serial numbers, who they're assigned "
         "to, and is it a replacement or new joiner?'. A blank assignee means it isn't "
         "assigned (in stock); a blank type means there's no Asset-History record."]
    if not vp["vendors"]:
        L.append("\nNo laptop purchase register is connected (the 'New Laptops "
                 "purchased' sheet has no rows).")
        return "\n".join(L)
    for v in vp["vendors"]:
        L.append(f"\n## {v['vendor']} — {v['count']} laptop(s)")
        L.append("| Serial | Laptop | Assigned to | New joiner / Replacement |")
        L.append("|---|---|---|---|")
        for it in v["laptops"]:
            L.append(f"| {it['serial']} | {it['laptop']} "
                     f"| {it['assignee'] or 'In stock / unassigned'} "
                     f"| {it['type'] or 'not recorded'} |")
    return "\n".join(L)


def build_bot_context(data: dict) -> str:
    """Full context for the IT Helper bot — everything an HR head asks about,
    beyond the aggregate report: per-person laptops, upcoming joiners + onboarding,
    returns/offboarding, peripherals, open tickets, software & budget."""
    return "\n\n".join([
        build_employee_directory(data),
        _bot_assignment_type_section(data),
        _bot_vendor_purchase_section(data),
        _bot_joiners_section(data),
        _bot_joiners_history_section(data),
        _bot_returns_section(data),
        _bot_peripherals_section(data),
        _bot_tickets_section(data),
        _bot_software_section(data),
        _bot_unplanned_section(data),
        _bot_delivery_section(data),
        _bot_sold_section(data),
        _bot_other_stock_section(data),
        _bot_backup_section(data),
        _bot_config_section(data),
    ])


def get_age_distribution(data: dict) -> dict:
    buckets = {"0-2yr": 0, "2-3yr": 0, "3-3.5yr": 0, "3.5-4yr": 0, ">4yr": 0}
    for row in data["assigned"]:
        if not is_truly_assigned(row):
            continue
        dt = parse_date(row.get("Warranty Start Date"))
        if not dt:
            continue
        yrs = age_years(dt)
        if yrs > 4:
            buckets[">4yr"] += 1
        elif yrs > 3.5:
            buckets["3.5-4yr"] += 1
        elif yrs > 3:
            buckets["3-3.5yr"] += 1
        elif yrs > 2:
            buckets["2-3yr"] += 1
        else:
            buckets["0-2yr"] += 1
    return buckets


def get_laptop_spend(data: dict) -> dict:
    """Extract laptop procurement spend from Actual Spends sheet.

    The sheet has columns like: Model, Joiners, Jan Joiners, Jan Spend,
    Feb Joiners, Feb Spend, March Joiners, Mar Spend, etc.
    Returns dict with keys: models (list of per-model data), total_joiners, total_spend.
    """
    MONTH_ABBREVS = {
        1: ["jan"], 2: ["feb"], 3: ["mar", "march"], 4: ["apr", "april"],
        5: ["may"], 6: ["jun", "june"], 7: ["jul", "july"], 8: ["aug"],
        9: ["sep", "sept"], 10: ["oct"], 11: ["nov"], 12: ["dec"],
    }
    abbrevs = MONTH_ABBREVS.get(TODAY.month, [])

    result = {"models": [], "total_joiners": 0, "units_this_month": 0,
              "total_spend": 0.0, "total_spend_inr": 0.0}
    total_row = None
    for row in data["actual_spend"]:
        # The row label (laptop model/category) is in the first column, whose
        # header is blank → read as 'col_0'. (The 'Joiners' column is the annual
        # total, NOT the label.)
        model = row.get("Model") or row.get("col_0", "")
        model_str = str(model).strip().lower() if model else ""
        # Capture the Total row separately for authoritative monthly spend
        if model_str in ("total", "grand total"):
            total_row = row
            continue
        if not model_str or model_str in ("none",):
            continue

        # Per-model: joiners (people) and units actually purchased this month.
        # The per-model "<month> Spend" cell is the QUANTITY of laptops bought
        # that month (the actual money lives only in the Total row). Laptops
        # purchased = sum of these per-model quantities, NOT the joiner counts.
        joiners = 0
        units = 0
        for key, val in row.items():
            key_lower = str(key).strip().lower()
            for abbr in abbrevs:
                if abbr not in key_lower:
                    continue
                num = _to_number(val)
                if "joiner" in key_lower:
                    joiners = int(num) if num is not None else 0
                elif "spend" in key_lower:
                    units = int(num) if num is not None else 0

        if joiners:
            result["total_joiners"] += joiners
        if units:
            result["models"].append({"model": str(model).strip(), "units": units})
            result["units_this_month"] += units

    # Authoritative figures from the Total row (the per-model rows are unreliable).
    # Spend is INR → convert to USD; joiner count is taken as-is.
    if total_row:
        for key, val in total_row.items():
            key_lower = str(key).strip().lower()
            for abbr in abbrevs:
                if abbr in key_lower and "spend" in key_lower:
                    num = _to_number(val)
                    result["total_spend"] = inr_to_usd(num) if num is not None else 0.0
                    result["total_spend_inr"] = num if num is not None else 0.0
                elif abbr in key_lower and "joiner" in key_lower:
                    num = _to_number(val)
                    if num is not None:
                        result["total_joiners"] = int(num)

    return result


# Month order for the financial year (Jan–Dec), abbr → display.
_FY_MONTHS = [("jan", "Jan"), ("feb", "Feb"), ("mar", "Mar"), ("apr", "Apr"),
              ("may", "May"), ("jun", "Jun"), ("jul", "Jul"), ("aug", "Aug"),
              ("sep", "Sep"), ("oct", "Oct"), ("nov", "Nov"), ("dec", "Dec")]


def get_laptop_spend_history(data: dict) -> dict:
    """Month-by-month laptop spend for the year, from the Actual Spends sheet:
    spend (₹) from the Total row, units (qty) summed from the per-model rows.
    Only includes months up to the current one. Returns {months, ytd_units,
    ytd_spend_inr}."""
    total_row = None
    model_rows = []
    for row in data.get("actual_spend", []):
        label = str(row.get("Model") or row.get("col_0", "") or "").strip().lower()
        if label in ("total", "grand total"):
            total_row = row
        elif label and label != "none":
            model_rows.append(row)

    months, ytd_units, ytd_spend = [], 0, 0.0
    for i, (abbr, disp) in enumerate(_FY_MONTHS):
        if i + 1 > TODAY.month:  # don't show future months
            break
        spend = None
        if total_row:
            for k, v in total_row.items():
                kl = str(k).strip().lower()
                if abbr in kl and "spend" in kl:
                    spend = _to_number(v)
        units = 0
        per_model = []
        for row in model_rows:
            label = str(row.get("Model") or row.get("col_0", "") or "").strip()
            for k, v in row.items():
                kl = str(k).strip().lower()
                if abbr in kl and "spend" in kl:
                    n = _to_number(v)
                    if n:
                        units += int(n)
                        per_model.append({"model": label, "units": int(n)})
        months.append({"month": disp, "units": units, "spend_inr": spend or 0.0,
                       "models": per_model})
        ytd_units += units
        ytd_spend += spend or 0.0
    return {"months": months, "ytd_units": ytd_units, "ytd_spend_inr": ytd_spend}


def get_recent_purchases(data: dict, days: int = 30) -> list[dict]:
    """Get laptops purchased within the given number of past days (up to today)."""
    cutoff = TODAY - timedelta(days=days)
    purchases = []
    for row in data["purchased"]:
        d = parse_date(row.get("Warranty Start Date"))
        if d and cutoff <= d <= TODAY:
            purchases.append({
                "brand": row.get("Brand", ""),
                "model": row.get("Model", ""),
                "serial": row.get("Serial no", ""),
                "date": d,
            })
    purchases.sort(key=lambda x: x["date"], reverse=True)
    return purchases


def get_purchases_this_month(data: dict) -> list[dict]:
    """Get laptops purchased within the current calendar month."""
    purchases = []
    for row in data["purchased"]:
        d = parse_date(row.get("Warranty Start Date"))
        if d and d.year == TODAY.year and d.month == TODAY.month and d <= TODAY:
            purchases.append({
                "brand": str(row.get("Brand", "") or "").strip(),
                "model": str(row.get("Model", "") or "").strip(),
                "serial": str(row.get("Serial no", "") or "").strip(),
                "configuration": str(row.get("Configuration", "") or "").strip(),
                "date": d,
            })
    purchases.sort(key=lambda x: x["date"], reverse=True)
    return purchases


def get_purchases_by_vendor(data: dict) -> dict:
    """Laptops grouped by the vendor they were bought from, from the 'New Laptops
    purchased' register's 'Purchased From' column. Lets the bot answer 'which
    vendor did we buy from?'. Returns {vendors:[{vendor,count,models:[{model,
    units}]}], total, date_from, date_to}."""
    by_vendor: dict[str, dict] = {}
    total = 0
    dmin = dmax = None
    for row in data.get("purchased", []):
        brand = str(row.get("Brand", "") or "").strip()
        model = str(row.get("Model", "") or "").strip()
        serial = str(row.get("Serial no", "") or "").strip()
        if not (brand or model or serial):
            continue  # skip blank/spacer rows
        vendor = str(row.get("Purchased From", "") or "").strip() or "Not recorded"
        label = (f"{brand} {model}").strip() or model or brand or "—"
        v = by_vendor.setdefault(vendor, {"count": 0, "models": {}})
        v["count"] += 1
        v["models"][label] = v["models"].get(label, 0) + 1
        total += 1
        d = parse_date(row.get("Warranty Start Date"))
        if d:
            dmin = d if dmin is None or d < dmin else dmin
            dmax = d if dmax is None or d > dmax else dmax
    vendors = []
    for name, v in by_vendor.items():
        models = [{"model": m, "units": u}
                  for m, u in sorted(v["models"].items(), key=lambda x: -x[1])]
        vendors.append({"vendor": name, "count": v["count"], "models": models})
    vendors.sort(key=lambda x: -x["count"])
    return {"vendors": vendors, "total": total, "date_from": dmin, "date_to": dmax}


# Row names in spend tracker that are laptop/hardware costs, not app subscriptions
# Match the specific hardware-procurement row name(s) — not loose substrings
# like "laptop", which could catch Microsoft Surface Laptop etc.
HARDWARE_SPEND_KEYWORDS = ["laptops procurement"]
# Row names that are aggregate/total rows (would double-count if summed)
TOTAL_ROW_KEYWORDS = ["total", "grand total", "sum"]


def _is_hardware_row(row: dict) -> bool:
    app_name = str(row.get("APPLICATION / SW / LICENSE", "")).lower()
    return any(kw in app_name for kw in HARDWARE_SPEND_KEYWORDS)


def _is_total_row(row: dict) -> bool:
    app_name = str(row.get("APPLICATION / SW / LICENSE", "")).strip().lower()
    return app_name in TOTAL_ROW_KEYWORDS


_warned_spend_keys: set = set()


def _warn_spend_once(key: str, message: str) -> None:
    if key not in _warned_spend_keys:
        _warned_spend_keys.add(key)
        print(f"  [spend] WARNING: {message}", file=sys.stderr)


def _spend_has_data(data: dict, key) -> bool:
    """True if any non-total row has a numeric value in the given month column."""
    return any(_to_number(r.get(key)) is not None for r in data["spend"] if not _is_total_row(r))


def _spend_period(data: dict) -> tuple:
    """Choose which month column to report spend from.

    Prefers the current month; if that column has no data entered yet, falls
    back to the most recent earlier month that does. Returns
    (key, date, is_current). When no month has data, returns the current
    month's column (so totals are 0) or (None, None, True) if no column matches.
    """
    cols = {}
    for row in data["spend"]:
        for key in row:
            d = parse_date(key)
            if d:
                cols[key] = d
    current = next(((k, d) for k, d in cols.items()
                    if d.year == TODAY.year and d.month == TODAY.month), None)
    if current and _spend_has_data(data, current[0]):
        return current[0], current[1], True
    # Fall back to the latest earlier month that has data.
    past = sorted((d, k) for k, d in cols.items() if d <= TODAY)
    for d, k in reversed(past):
        if _spend_has_data(data, k):
            return k, d, False
    if current:
        return current[0], current[1], True
    return None, None, True


def _spend_asof_note(data: dict) -> str:
    """' (as of May)' when spend is reported from a fallback (earlier) month
    because the current month has no data yet; '' when it's the current month."""
    _, d, is_current = _spend_period(data)
    return "" if is_current or d is None else f" _(as of {d.strftime('%b')})_"


def _spend_month_label(data: dict) -> str:
    """Reporting-month name for the spend section, e.g. 'June' or
    'May (latest with data)' when the current month isn't filled in yet."""
    _, d, is_current = _spend_period(data)
    label = (d or TODAY).strftime('%B')
    return label if is_current else f"{label} (latest with data)"


def get_software_spend_this_month(data: dict) -> float:
    """Total software / subscription / license spend for the reporting month.

    This is the spend tracker's monthly total — the sum of every named line
    item in that month's column. Rows with a blank APPLICATION name (e.g. the
    sheet's own green Total row, which has no name) and explicit total rows are
    skipped, so the figure equals the sheet total without double counting.
    """
    month_key, _d, is_current = _spend_period(data)
    total = 0.0
    line_items = 0
    if month_key:
        for row in data["spend"]:
            name = str(row.get("APPLICATION / SW / LICENSE", "") or "").strip()
            if not name or _is_total_row(row):
                continue
            val = _to_number(row.get(month_key))
            if val is not None:
                total += val
                line_items += 1
    _spend_diagnostic_once(data, month_key, is_current, total, line_items)
    return total


def get_software_spend_history(data: dict) -> dict:
    """Month-by-month software/subscription spend this year (sum of named line
    items per month column), up to the current month. {months, ytd}."""
    cols = {}
    for row in data["spend"]:
        for k in row:
            d = parse_date(k)
            if d:
                cols[k] = d
    months, ytd = [], 0.0
    for key, d in sorted(cols.items(), key=lambda kv: kv[1]):
        if d.year != TODAY.year or d.month > TODAY.month:
            continue
        total = 0.0
        for row in data["spend"]:
            name = str(row.get("APPLICATION / SW / LICENSE", "") or "").strip()
            if not name or _is_total_row(row):
                continue
            v = _to_number(row.get(key))
            if v is not None:
                total += v
        months.append({"month": d.strftime("%b"), "spend": total})
        ytd += total
    return {"months": months, "ytd": ytd}


def get_joiners_history(data: dict) -> dict:
    """People who joined each month this year (by Confirm DOJ), up to the
    current month, with their names. {months:[{month,count,names}], ytd}."""
    by_month: dict[int, list] = defaultdict(list)
    for row in data.get("joinings", []):
        d = parse_date(row.get("Confirm DOJ") or row.get("DOJ As per Offer letter"))
        if d and d.year == TODAY.year and d <= TODAY:
            name = str(row.get("Employee name") or row.get("Employee Name")
                       or row.get("Name") or "").strip()
            by_month[d.month].append(name or "Unknown")
    months, ytd = [], 0
    for m in range(1, TODAY.month + 1):
        names = by_month.get(m, [])
        months.append({"month": dt.date(TODAY.year, m, 1).strftime("%b"),
                       "count": len(names), "names": names})
        ytd += len(names)
    return {"months": months, "ytd": ytd}


def get_assignment_history(data: dict) -> dict:
    """Laptop assignments per month this year, split into new-joiner vs
    replacement, from the Asset History sheet (Assigned Date + the
    'New Joiner/Replacement' column). {months:[{month,new_joiner,replacement}],
    ytd_new_joiner, ytd_replacement}."""
    by_month: dict[int, dict] = defaultdict(lambda: {"new_joiner": 0, "replacement": 0})
    for row in data.get("history", []):
        d = parse_date(row.get("Assigned Date"))
        if not d or d.year != TODAY.year or d > TODAY:
            continue
        typ = str(row.get("New Joiner/Replacement", "") or "").strip().lower()
        if "replace" in typ:
            by_month[d.month]["replacement"] += 1
        elif "joiner" in typ or "new" in typ:
            by_month[d.month]["new_joiner"] += 1
    months, nj, rep = [], 0, 0
    for m in range(1, TODAY.month + 1):
        b = by_month.get(m, {"new_joiner": 0, "replacement": 0})
        months.append({"month": dt.date(TODAY.year, m, 1).strftime("%b"),
                       "new_joiner": b["new_joiner"], "replacement": b["replacement"]})
        nj += b["new_joiner"]
        rep += b["replacement"]
    return {"months": months, "ytd_new_joiner": nj, "ytd_replacement": rep}


def get_software_inventory(data: dict) -> list[dict]:
    """Per-app/subscription line items for the reporting month, with renewal
    dates — the software & licence inventory."""
    month_key, _d, _ = _spend_period(data)
    items = []
    for row in data["spend"]:
        name = str(row.get("APPLICATION / SW / LICENSE", "") or "").strip()
        if not name or _is_total_row(row):
            continue
        items.append({
            "app": name,
            "frequency": str(row.get("FREQUENCY", "") or "").strip(),
            "payment": str(row.get("Payment Method", "") or "").strip(),
            "dept": str(row.get("Department", "") or "").strip(),
            "cost": _to_number(row.get(month_key)) if month_key else None,
            "renewal": parse_date(row.get("Renewal data") or row.get("Renewal Date")),
        })
    items.sort(key=lambda x: (x["cost"] is None, -(x["cost"] or 0)))
    return items


def get_upcoming_renewals(data: dict, days: int = 60) -> list[dict]:
    """Subscriptions whose renewal date falls within the next `days`."""
    cutoff = TODAY + timedelta(days=days)
    ren = [i for i in get_software_inventory(data)
           if i["renewal"] and TODAY <= i["renewal"] <= cutoff]
    ren.sort(key=lambda x: x["renewal"])
    return ren


def get_budget_vs_actual(data: dict) -> dict:
    """IT budget vs actual: laptop procurement (annual plan vs spend), how many
    laptops were procured this month, plus the current month's software spend."""
    pace = get_spend_pace(data)
    ls = get_laptop_spend(data)
    return {
        "laptop_annual_budget": pace["annual_planned"],
        "laptop_monthly_budget": pace["monthly_planned"],
        "laptop_annual_inr": pace["annual_planned_inr"],
        "laptop_monthly_inr": pace["monthly_planned_inr"],
        "laptop_actual_this_month": pace["actual"],
        "laptop_pct_of_monthly": pace["pct_used"],
        "laptops_this_month": ls.get("units_this_month", 0),
        "laptop_models": ls.get("models", []),
        "laptop_spend_inr": ls.get("total_spend_inr", 0.0),
        "spend_history": get_laptop_spend_history(data),
        "purchases_this_month": get_purchases_this_month(data),
        "new_laptops_registered": len(get_purchases_this_month(data)),
        "software_this_month": get_software_spend_this_month(data),
    }


def _spend_diagnostic_once(data, month_key, is_current, software_total, line_items) -> None:
    """One-time stderr dump of how the spend sheet was parsed, so figure
    mismatches can be traced from the Actions log."""
    if "spend-debug" in _warned_spend_keys:
        return
    _warned_spend_keys.add("spend-debug")
    cols = {}
    all_keys: set = set()
    for row in data["spend"]:
        all_keys.update(row.keys())
        for k in row:
            d = parse_date(k)
            if d:
                cols.setdefault(k, d)
    detected = ", ".join(f"{k!r}→{d.strftime('%b %Y')}"
                         for k, d in sorted(cols.items(), key=lambda x: x[1]))
    unparsed = [k for k in all_keys if k not in cols]
    print(f"  [spend] columns: {sorted(map(str, all_keys))}", file=sys.stderr)
    print(f"  [spend] detected month columns: {detected or '(none)'}", file=sys.stderr)
    print(f"  [spend] NOT parsed as months: {sorted(map(str, unparsed))}", file=sys.stderr)
    print(f"  [spend] reporting month={month_key!r} is_current={is_current} "
          f"→ software total {software_total:.2f} from {line_items} line item(s)",
          file=sys.stderr)


def get_current_month_spend(data: dict) -> tuple[float, list[dict], float, float]:
    """Get app spend for the current month (or latest month with data), upcoming
    renewals, hardware spend, and grand total.

    Returns: (app_only_total, renewals, hardware_total, grand_total)
        - app_only_total: sum of app/subscription rows (excludes hardware + Total rows)
        - hardware_total: sum of "Laptops Procurement", "Antivirus,MDM" type rows
        - grand_total: app + hardware (matches the Total row in the sheet)
    """
    month_key, _month_date, is_current = _spend_period(data)

    app_total = 0.0
    hw_total = 0.0
    numeric_cells = 0
    if month_key:
        for row in data["spend"]:
            if _is_total_row(row):
                continue
            val = _to_number(row.get(month_key))
            if val is None:
                continue
            numeric_cells += 1
            if _is_hardware_row(row):
                hw_total += val
            else:
                app_total += val

    # Warn (once) only when no month anywhere has spend data — the fallback in
    # _spend_period already handles a not-yet-filled current month.
    if month_key is None or app_total == 0:
        _warn_spend_once("no-data", "no spend data found in any month column of the spend sheet")

    # Upcoming renewals (app only)
    renewals = []
    cutoff = TODAY + timedelta(days=30)
    for row in data["spend"]:
        if _is_hardware_row(row) or _is_total_row(row):
            continue
        rd = parse_date(row.get("Renewal data"))
        if rd and TODAY <= rd <= cutoff:
            renewals.append({
                "app": row.get("APPLICATION / SW / LICENSE", "Unknown"),
                "date": rd,
                "dept": row.get("Department", ""),
                "frequency": row.get("FREQUENCY", ""),
            })
    renewals.sort(key=lambda x: x["date"])
    return app_total, renewals, hw_total, app_total + hw_total


def _month_column_key(data: dict, year: int, month: int):
    """Return the spend-sheet column key matching the given year/month, or None."""
    for row in data["spend"]:
        for key in row:
            d = parse_date(key)
            if d and d.year == year and d.month == month:
                return key
    return None


def get_app_spend_detail(data: dict) -> dict:
    """Richer app-spend breakdown for the monthly report.

    Returns: this_month, last_month (USD totals, app rows only), delta,
    top_apps (list of {app, dept, amount}), by_dept (dict dept→amount).
    """
    # Use the same reporting month as get_current_month_spend (falls back to the
    # latest month with data when the current one is empty) so figures match.
    this_key, this_date, _ = _spend_period(data)
    if this_date:
        prev_dt = this_date.replace(day=1) - timedelta(days=1)
        prev_key = _month_column_key(data, prev_dt.year, prev_dt.month)
    else:
        prev_key = None

    this_total = 0.0
    last_total = 0.0
    per_app = []
    by_dept: dict[str, float] = defaultdict(float)
    for row in data["spend"]:
        if _is_hardware_row(row) or _is_total_row(row):
            continue
        app = str(row.get("APPLICATION / SW / LICENSE", "")).strip()
        dept = str(row.get("Department", "")).strip() or "Unassigned"
        tv = _to_number(row.get(this_key)) if this_key else None
        pv = _to_number(row.get(prev_key)) if prev_key else None
        if tv is not None:
            this_total += tv
            if app:
                per_app.append({"app": app, "dept": dept, "amount": tv})
                by_dept[dept] += tv
        if pv is not None:
            last_total += pv

    per_app.sort(key=lambda x: x["amount"], reverse=True)
    return {
        "this_month": this_total,
        "last_month": last_total,
        "delta": this_total - last_total,
        "top_apps": per_app[:5],
        "by_dept": dict(sorted(by_dept.items(), key=lambda kv: kv[1], reverse=True)),
    }


def get_upcoming_joiners(data: dict, days: int = 14) -> list[dict]:
    cutoff = TODAY + timedelta(days=days)
    joiners = []
    for row in data["joinings"]:
        dt = parse_date(row.get("Confirm DOJ") or row.get("DOJ As per Offer letter"))
        if dt and TODAY <= dt <= cutoff:
            joiners.append({
                "name": row.get("Employee name", "Unknown"),
                "doj": dt,
                "designation": row.get("Designation", ""),
                "department": row.get("Department", ""),
            })
    joiners.sort(key=lambda x: x["doj"])
    return joiners


def lookup_laptop_config(data: dict, department: str, designation: str) -> str:
    """Find the standard laptop config for a joiner based on department/designation.

    Returns a short string like "Lenovo L14 / 16GB / i5" or "" if no match.
    """
    if not department and not designation:
        return ""
    dept_l = str(department).strip().lower()
    desig_l = str(designation).strip().lower()
    best_match = None
    for row in data.get("configuration", []):
        cfg_dept = str(row.get("Department & Owner", "")).strip().lower()
        cfg_role = str(row.get("Role / Position", "")).strip().lower()
        # Match if department contains cfg_dept or vice versa, or role matches
        dept_hit = cfg_dept and (cfg_dept in dept_l or dept_l in cfg_dept)
        role_hit = cfg_role and (cfg_role in desig_l or desig_l in cfg_role)
        if dept_hit and role_hit:
            best_match = row
            break  # exact match on both
        if dept_hit and not best_match:
            best_match = row
    if not best_match:
        return ""
    device = str(best_match.get("Device Type", "")).strip()
    ram = str(best_match.get("RAM", "")).strip()
    proc = str(best_match.get("Processor", "")).strip()
    parts = [p for p in (device, ram, proc) if p and p.lower() != "none"]
    return " / ".join(parts)


def get_joiners_with_laptop_needs(data: dict, days: int = 7) -> list[dict]:
    """Upcoming joiners in the next N days with their required laptop config."""
    joiners = get_upcoming_joiners(data, days)
    for j in joiners:
        j["laptop_config"] = lookup_laptop_config(data, j["department"], j["designation"])
        j["days_until"] = (j["doj"] - TODAY).days
    return joiners


ONBOARDING_CHECKLIST_COLS = [
    "Email ID Creation",
    "Reporting Manager Update",
    "Enable MFA",
    "Invite on Clickup",
    "Invite on slack",
    "Asset policy Acknowledgement",
]


def get_onboarding_readiness(data: dict, days: int = 7) -> Optional[float]:
    """Average % of checklist items complete for joiners DOJ in next N days.

    Returns None if there are no joiners in the window or no matching
    checklist rows.
    """
    joiners = get_upcoming_joiners(data, days)
    if not joiners:
        return None
    checklist_rows = data.get("checklist", [])
    if not checklist_rows:
        return None

    def _norm(s) -> str:
        return str(s or "").strip().lower()

    by_name = {}
    for row in checklist_rows:
        name = row.get("Name ", row.get("Name", ""))
        if name:
            by_name[_norm(name)] = row

    pct_values: list[float] = []
    for j in joiners:
        row = by_name.get(_norm(j.get("name")))
        if not row:
            continue
        done = 0
        total = 0
        for col in ONBOARDING_CHECKLIST_COLS:
            v = row.get(col)
            if v is None:
                continue
            total += 1
            if str(v).strip().lower() not in ("", "none", "no", "pending", "0"):
                done += 1
        if total:
            pct_values.append(done / total * 100)
    if not pct_values:
        return None
    return sum(pct_values) / len(pct_values)


def get_cost_per_joiner(data: dict) -> Optional[float]:
    """Laptop spend MTD / joiners-with-spend-this-month."""
    ls = get_laptop_spend(data)
    if ls["total_spend"] > 0 and ls["total_joiners"] > 0:
        return ls["total_spend"] / ls["total_joiners"]
    return None


def get_stock_vs_joiners(data: dict, days: int = 7) -> dict:
    """Compare current laptop stock to upcoming joiner demand.

    Returns a dict with keys:
      stock_ready, stock_backup, joiners_next_week, joiners_next_30_days,
      gap_next_week, gap_next_30_days, status ("ok"/"watch"/"short")
    """
    stock_ready = len(data["in_stock"])
    stock_backup = len(data["backup"])
    joiners_7 = get_upcoming_joiners(data, days)
    joiners_30 = get_upcoming_joiners(data, 30)
    gap_7 = len(joiners_7) - stock_ready
    gap_30 = len(joiners_30) - stock_ready
    if gap_7 > 0:
        status = "short"
    elif gap_30 > 0:
        status = "watch"
    else:
        status = "ok"
    return {
        "stock_ready": stock_ready,
        "stock_backup": stock_backup,
        "joiners_next_week": len(joiners_7),
        "joiners_next_30_days": len(joiners_30),
        "gap_next_week": gap_7,
        "gap_next_30_days": gap_30,
        "status": status,
    }


# ---------------------------------------------------------------------------
# IT issues (helpdesk tickets)
# ---------------------------------------------------------------------------

# Flexible header matching — the eventual ticketing-tool / email / Slack
# integration may label columns slightly differently.
_ISSUE_FIELDS = {
    "date": ("date raised", "raised date", "date", "created", "reported on"),
    "issue": ("issue", "subject", "summary", "title", "description", "problem"),
    "raised_by": ("raised by", "reported by", "requester", "employee", "user"),
    "priority": ("priority", "severity", "urgency"),
    "status": ("status", "state"),
    "owner": ("owner", "assigned to", "assignee", "handled by"),
    "latest_update": ("latest update", "latest comment", "activity", "update note", "work note"),
}

OPEN_STATUSES = ("open", "new", "in progress", "in-progress", "pending", "on hold", "reopened",
                 "to do", "todo", "blocked", "waiting", "acknowledged", "assigned")
# A ticket is considered resolved only when its status is clearly terminal.
# Anything else (including unrecognised custom statuses) counts as open, which
# is the conservative, actionable default for an IT manager.
CLOSED_STATUSES = ("resolved", "closed", "done", "complete", "completed",
                   "cancelled", "canceled", "wont do", "won't do", "duplicate", "rejected")


def _match_issue_field(row: dict, field: str):
    wanted = _ISSUE_FIELDS[field]
    for key, val in row.items():
        k = str(key).strip().lower()
        if any(w == k or w in k for w in wanted):
            return val
    return None


def _issue_pending_remark(issue: dict) -> str:
    """The 'why is this pending' remark for an IT ticket.

    Prefers the latest activity/comment on the ticket (the real explanation,
    e.g. "waiting on employee to visit Lenovo service center"). Falls back to a
    status/age-derived remark only when there is no comment. Resolved tickets
    get their closing status."""
    if not issue["is_open"]:
        return issue["status"] or "Resolved"

    latest = (issue.get("latest_update") or "").strip()
    if latest:
        return latest

    s = issue["status"].lower()
    if "progress" in s:
        why = "Being worked on"
    elif "due today" in s:
        why = "Due today — needs closing"
    elif "overdue" in s:
        why = "Overdue — escalate"
    elif "hold" in s or "wait" in s or "block" in s:
        why = "Blocked / awaiting input"
    elif s in ("to do", "todo", "open", "new", "reopened", ""):
        why = "Not started — awaiting pickup"
    else:
        why = issue["status"] or "Open"

    parts = [why]
    if issue["date"]:
        days = (TODAY - issue["date"]).days
        parts.append("raised today" if days <= 0 else f"open {days}d")
    if issue["priority"].lower() in ("high", "critical", "urgent", "p1"):
        parts.append("high priority")
    if not issue["owner"]:
        parts.append("unassigned")
    return " · ".join(parts)


def get_it_issues(data: dict) -> dict:
    """Summarise IT helpdesk issues from the optional 'IT Issues' sheet.

    Returns a dict with keys:
      connected (bool) — whether any issue source/data is present
      issues (list)    — normalised rows
      open, resolved, high_open (int)

    When no source is connected, `connected` is False and the report shows a
    placeholder rather than fabricating tickets.
    """
    rows = data.get("it_issues", [])
    issues = []
    for row in rows:
        issue = _match_issue_field(row, "issue")
        status_raw = _match_issue_field(row, "status")
        if not issue and not status_raw:
            continue
        status = str(status_raw or "").strip()
        priority = str(_match_issue_field(row, "priority") or "").strip()
        rec = {
            "date": parse_date(_match_issue_field(row, "date")),
            "issue": str(issue or "").strip(),
            "raised_by": str(_match_issue_field(row, "raised_by") or "").strip(),
            "priority": priority,
            "status": status,
            "owner": str(_match_issue_field(row, "owner") or "").strip(),
            "latest_update": str(_match_issue_field(row, "latest_update") or "").strip(),
            # Resolved only when the status is clearly terminal; everything else
            # (incl. blank or unknown custom statuses) is treated as open.
            "is_open": status.lower() not in CLOSED_STATUSES,
        }
        rec["remark"] = _issue_pending_remark(rec)
        issues.append(rec)
    open_issues = [i for i in issues if i["is_open"]]
    high_open = [i for i in open_issues if i["priority"].lower() in ("high", "critical", "urgent", "p1")]
    # Sort: open first, then high priority, then most recent
    issues.sort(key=lambda i: (not i["is_open"], i["date"] or dt.date.min), reverse=False)
    return {
        "connected": bool(rows),
        "issues": issues,
        "open": len(open_issues),
        "resolved": len(issues) - len(open_issues),
        "high_open": len(high_open),
    }


# ---------------------------------------------------------------------------
# Action items for the IT manager
# ---------------------------------------------------------------------------

def get_action_items(data: dict, hc: dict) -> list[dict]:
    """Synthesise a prioritised, deduplicated to-do list for the IT manager
    from the signals already in the data. Each item is {priority, text} where
    priority is one of 🔴/🟡/🟢 (sorted most urgent first)."""
    items: list[dict] = []

    runway = hc["_runway_weeks"]
    joiners_7 = hc["_joiners_7"]
    stock_ready = hc["_stock_ready"]
    critical = hc["_critical_aging"]
    pace_pct = hc["_pace_pct"]

    # 1. Joiner stock shortage (most time-critical)
    if joiners_7 > stock_ready:
        gap = joiners_7 - stock_ready
        items.append({"priority": "🔴",
                      "text": f"Order {gap} laptop(s) now — {joiners_7} joiners next week vs {stock_ready} in stock"})
    elif joiners_7 and stock_ready - joiners_7 <= 2:
        items.append({"priority": "🟡",
                      "text": f"Tight stock for joiners ({stock_ready} ready vs {joiners_7} next week) — line up procurement"})

    # 2. Stock runway
    if runway is not None and runway < 2:
        items.append({"priority": "🔴",
                      "text": f"Replenish laptop stock — runway only {runway} weeks at current pace"})
    elif runway is not None and runway < 4:
        items.append({"priority": "🟡",
                      "text": f"Plan a laptop purchase — runway {runway} weeks"})

    # 3. Critical aging replacements
    if critical > 0:
        pr = "🔴" if critical > 5 else "🟡"
        items.append({"priority": pr,
                      "text": f"Schedule replacement of {critical} laptop(s) over 4 years old"})

    # 4. Onboarding gaps for joiners next 7 days
    readiness = get_onboarding_readiness(data, 7)
    if readiness is not None and readiness < 80:
        items.append({"priority": "🟡",
                      "text": f"Close onboarding gaps — checklist only {readiness:.0f}% complete for next week's joiners"})

    # 5. Upcoming renewals
    _, renewals, _, _ = get_current_month_spend(data)
    if renewals:
        items.append({"priority": "🟡",
                      "text": f"Review {len(renewals)} subscription renewal(s) due in the next 30 days"})

    # 6. Spend pace
    if pace_pct is not None and pace_pct > 110:
        items.append({"priority": "🔴",
                      "text": f"Laptop spend at {pace_pct:.0f}% of monthly budget — review procurement spend"})

    # 7. Open high-priority IT issues
    issues = get_it_issues(data)
    if issues["high_open"]:
        items.append({"priority": "🔴",
                      "text": f"Resolve {issues['high_open']} high-priority IT issue(s) still open"})
    elif issues["open"]:
        items.append({"priority": "🟢",
                      "text": f"Work down {issues['open']} open IT issue(s)"})

    if not items:
        items.append({"priority": "🟢", "text": "No action items — stock, aging, spend and onboarding all healthy"})

    order = {"🔴": 0, "🟡": 1, "🟢": 2}
    items.sort(key=lambda x: order.get(x["priority"], 3))
    return items


# ---------------------------------------------------------------------------
# Report generators
# ---------------------------------------------------------------------------

def _overall_status(hc: dict) -> str:
    statuses = (hc["stock"], hc["aging"], hc["joiner_prep"], hc["spend"])
    if "🔴" in statuses:
        return "🔴 Action Needed"
    if "🟡" in statuses:
        return "🟡 Needs Attention"
    return "🟢 On Track"


def _short_model(make, model) -> str:
    """Tidy a make/model for one-line display (drop the trailing 'Laptop')."""
    name = f"{make} {model}".strip()
    if name.lower().endswith(" laptop"):
        name = name[:-len(" laptop")]
    return name


def _aging_slack_line(a: dict) -> str:
    """One clean aging line: name — model · age, with a warranty note only when
    it's actually actionable (expired)."""
    note = ""
    we = a.get("warranty_end")
    if we and we < TODAY:
        note = f" · _warranty expired {we.strftime('%b %Y')}_"
    return f"{a['employee']} — {_short_model(a['make'], a['model'])} · {a['age_years']}yr{note}"


def _truncate(text: str, n: int = 80) -> str:
    text = " ".join(str(text).split())
    return text if len(text) <= n else text[: n - 1].rstrip() + "…"


def build_report_slack(data: dict, prev_snap: Optional[dict], period: str) -> str:
    """The IT report as a Slack post — exactly the seven sections requested:
    open tickets, stock ready by OS, procurement, joiners, spend, aging, vendor
    payments."""
    issues = get_it_issues(data)
    open_issues = [i for i in issues["issues"] if i["is_open"]]
    stock_os = get_stock_by_os(data)
    stock_ready = len(data["in_stock"])
    aging = get_aging_laptops(data)
    critical = [a for a in aging if a["priority"] == "Critical"]
    joiners = get_joiners_with_laptop_needs(data, 30)
    laptop_spend = get_laptop_spend(data)
    software_total = get_software_spend_this_month(data)
    runway = get_procurement_runway(data)
    vendor = get_vendor_payments(data)

    L = [f"*📋 IT {period} Report — {TODAY.strftime('%d %b %Y')}*"]

    # 1) Open IT tickets (all of them)
    L.append(f"\n*1) 🐞 Open IT Tickets — {len(open_issues)}*")
    if not issues["connected"]:
        L.append("_No ticket source connected_")
    elif not open_issues:
        L.append("None open ✅")
    else:
        for i in open_issues:
            owner = i["owner"] or "unassigned"
            requester = i["raised_by"] or "unknown"
            L.append(f"› *{_truncate(i['issue'], 55)}* — _{_truncate(i['remark'], 90)}_ "
                     f"({i['status'] or 'Open'} · raised by {requester} · owner {owner})")

    # 2) Laptop stock ready — by OS, with configurations
    L.append(f"\n*2) 💻 Laptop Stock Ready — {stock_ready}*")
    if stock_ready == 0:
        L.append("None in ready stock")
    else:
        for os_label, items in stock_os.items():
            L.append(f"*{os_label} — {len(items)}*")
            for it in items:
                L.append(f"› {it['config']}")

    # 3) Procurement suggestion
    need = len(joiners) + len(critical) - stock_ready
    L.append("\n*3) 🛒 Procurement Suggestion*")
    if need > 0:
        L.append(f"⚠️ Order *{need}* laptop(s) — {len(joiners)} joiners (30d) + "
                 f"{len(critical)} critical replacements vs {stock_ready} ready")
    else:
        L.append(f"✅ Stock covers demand — {stock_ready} ready vs {len(joiners)} joiners "
                 f"+ {len(critical)} critical replacements")
    if runway["weeks"] is not None:
        L.append(f"_Runway ~{runway['weeks']} wks at {runway['avg_per_week']} joiners/wk_")

    # 4) Upcoming joiners (next 30 days)
    L.append(f"\n*4) ⏰ Upcoming Joiners (30d) — {len(joiners)}*")
    if not joiners:
        L.append("None in the next 30 days")
    else:
        for j in joiners:
            days = f"in {j['days_until']}d" if j['days_until'] > 0 else "today"
            cfg = f" · _{j['laptop_config']}_" if j['laptop_config'] else ""
            L.append(f"› {j['name']} — {j['department']} · DOJ {j['doj'].strftime('%d %b')} ({days}){cfg}")

    # 5) Spend this month — total software/license spend + laptop procurement
    L.append(f"\n*5) 💰 Spend This Month — {_spend_month_label(data)}*")
    L.append(f"Apps & licenses *{fmt_usd(software_total)}*")
    L.append(f"Laptops (procurement plan) *{fmt_usd(laptop_spend['total_spend'])}*")

    # 6) Laptop aging + what to do
    L.append(f"\n*6) ⏳ Laptop Aging — {len(aging)} over 3.5yr ({len(critical)} critical)*")
    if not aging:
        L.append("None over 3.5 years ✅")
    else:
        for a in aging[:8]:
            L.append(f"› {a['employee']} — {_short_model(a['make'], a['model'])} · "
                     f"{a['age_years']}yr → _{aging_action(a)}_")
        if len(aging) > 8:
            L.append(f"_+{len(aging) - 8} more (oldest shown first)_")

    # 7) Vendor payments pending (native INR)
    L.append(f"\n*7) 🧾 Vendor Payments Pending — {vendor['count']} · {fmt_inr_full(vendor['total_inr'])}*")
    if not vendor["connected"]:
        L.append("_No vendor payments sheet connected_")
    elif not vendor["pending"]:
        L.append("None pending ✅")
    else:
        for v in vendor["pending"]:
            due = v["due"].strftime('%d %b %Y') if v["due"] else "—"
            inv = f" ({v['invoice']})" if v["invoice"] else ""
            ov = f" · overdue {v['overdue']}" if v["overdue"] else ""
            L.append(f"› {v['vendor']}{inv} — *{fmt_inr_full(v['amount_inr'])}* · due {due}{ov}")

    L.append("\n_React 👍/👎 or reply with feedback._")
    return "\n".join(L)


def generate_weekly_slack(data: dict, prev_snap: Optional[dict] = None) -> str:
    return build_report_slack(data, prev_snap, "Weekly")


def generate_joiner_alert(data: dict) -> str:
    """Separate Slack alert: joiners next week + stock vs joiners analysis."""
    lines = [f"*🚨 Joiner Alert & Stock Check — {TODAY.strftime('%d %B %Y')}*\n"]

    # Joiners next week with laptop config
    joiners_week = get_joiners_with_laptop_needs(data, 7)
    lines.append(f"*⏰ Joiners Next Week* ({len(joiners_week)}) — prepare laptops")
    if joiners_week:
        for j in joiners_week:
            cfg = f" · _{j['laptop_config']}_" if j['laptop_config'] else ""
            days = f"in {j['days_until']}d" if j['days_until'] > 0 else "today"
            lines.append(f"• {j['name']} — {j['department']}, {j['designation']} (DOJ {j['doj'].strftime('%d %b')}, {days}){cfg}")
    else:
        lines.append("• No joiners in the next 7 days ✅")

    # Stock vs Joiners
    svj = get_stock_vs_joiners(data, 7)
    icon = {"ok": "✅", "watch": "🟡", "short": "🔴"}[svj["status"]]
    lines.append(f"\n*{icon} Stock vs Joiners*")
    lines.append(f"• Laptops in stock: {svj['stock_ready']} (ready) + {svj['stock_backup']} (backup)")
    lines.append(f"• Joiners next 7 days: {svj['joiners_next_week']} | next 30 days: {svj['joiners_next_30_days']}")
    if svj["gap_next_week"] > 0:
        lines.append(f"• 🔴 *Short {svj['gap_next_week']} laptop(s)* for next week's joiners — arrange immediately!")
    elif svj["gap_next_30_days"] > 0:
        lines.append(f"• 🟡 Will be short {svj['gap_next_30_days']} laptop(s) within 30 days — plan procurement")
    else:
        lines.append(f"• ✅ Stock covers next 30 days of joiners")

    return "\n".join(lines)


def _fx_footnote() -> str:
    rate = round(1 / INR_TO_USD_RATE, 1) if INR_TO_USD_RATE else 0
    return (f"_Currency: laptop procurement & vendor amounts are shown in ₹ (INR); "
            f"app/software subscriptions in $ (USD). Reference rate $1 ≈ ₹{rate} "
            f"(INR_TO_USD_RATE)._")


def build_report_full(data: dict, prev_snap: Optional[dict], period: str) -> str:
    """The IT report as a Markdown document — the same seven sections as the
    Slack post, with full tables. Saved to output/full-report.md."""
    issues = get_it_issues(data)
    open_issues = [i for i in issues["issues"] if i["is_open"]]
    stock_os = get_stock_by_os(data)
    stock_ready = len(data["in_stock"])
    aging = get_aging_laptops(data)
    critical = [a for a in aging if a["priority"] == "Critical"]
    joiners = get_joiners_with_laptop_needs(data, 30)
    laptop_spend = get_laptop_spend(data)
    software_total = get_software_spend_this_month(data)
    runway = get_procurement_runway(data)
    vendor = get_vendor_payments(data)

    L = [f"# IT {period} Report — {TODAY.strftime('%d %B %Y')}\n"]

    # 1. Open IT tickets
    L.append(f"## 1. Open IT Tickets — {len(open_issues)}\n")
    if not issues["connected"]:
        L.append("_No ticket source connected._")
    elif not open_issues:
        L.append("None open. ✅")
    else:
        L.append("| Issue | Status | Raised By | Owner | Why pending |")
        L.append("|-------|--------|-----------|-------|-------------|")
        for i in open_issues:
            L.append(f"| {i['issue']} | {i['status'] or 'Open'} | {i['raised_by'] or '—'} "
                     f"| {i['owner'] or '—'} | {i['remark']} |")

    # 2. Laptop stock ready — by OS, with configurations
    L.append(f"\n## 2. Laptop Stock Ready — {stock_ready}\n")
    if stock_ready == 0:
        L.append("None in ready stock.")
    else:
        L.append("| OS | Count |")
        L.append("|----|-------|")
        for os_label, items in stock_os.items():
            L.append(f"| {os_label} | {len(items)} |")
        L.append("\n**Configurations**\n")
        L.append("| OS | Make / Model | RAM | Processor | Asset Tag |")
        L.append("|----|--------------|-----|-----------|-----------|")
        for os_label, items in stock_os.items():
            for it in items:
                L.append(f"| {os_label} | {it['make']} {it['model']} | {it['ram'] or '—'} "
                         f"| {it['processor'] or '—'} | {it['tag'] or '—'} |")

    # 3. Procurement suggestion
    need = len(joiners) + len(critical) - stock_ready
    L.append("\n## 3. Procurement Suggestion\n")
    L.append(f"- Ready stock: **{stock_ready}**")
    L.append(f"- Demand (next 30 days): **{len(joiners)}** joiners + **{len(critical)}** critical replacements")
    if runway["weeks"] is not None:
        L.append(f"- Runway: ~{runway['weeks']} weeks at {runway['avg_per_week']} joiners/wk")
    if need > 0:
        L.append(f"\n**⚠️ Order {need} laptop(s)** to cover next-30-day demand.")
    else:
        L.append("\n**✅ No immediate purchase needed** — stock covers next-30-day demand.")

    # 4. Upcoming joiners
    L.append(f"\n## 4. Upcoming Joiners (next 30 days) — {len(joiners)}\n")
    if not joiners:
        L.append("None in the next 30 days.")
    else:
        L.append("| Name | Department | Designation | DOJ | Days | Laptop Config |")
        L.append("|------|------------|-------------|-----|------|---------------|")
        for j in joiners:
            days = f"{j['days_until']}d" if j['days_until'] > 0 else "today"
            L.append(f"| {j['name']} | {j['department']} | {j['designation']} "
                     f"| {j['doj'].strftime('%d %b %Y')} | {days} | {j['laptop_config'] or '—'} |")

    # 5. Spend this month — total software/license spend + laptop procurement
    L.append(f"\n## 5. Spend This Month — {_spend_month_label(data)}\n")
    L.append("| Category | Amount |")
    L.append("|----------|--------|")
    L.append(f"| **Apps & licenses** (all subscriptions) | **{fmt_usd(software_total)}** |")
    L.append(f"| Laptops (procurement plan) | {fmt_usd(laptop_spend['total_spend'])} |")

    # 6. Laptop aging + action
    L.append(f"\n## 6. Laptop Aging — {len(aging)} over 3.5yr ({len(critical)} critical)\n")
    if not aging:
        L.append("None over 3.5 years. ✅")
    else:
        L.append("| Employee | Dept | Make/Model | Age (yrs) | Priority | Action |")
        L.append("|----------|------|-----------|-----------|----------|--------|")
        for a in aging:
            L.append(f"| {a['employee']} | {a['department']} | {a['make']} {a['model']} "
                     f"| {a['age_years']} | {a['priority']} | {aging_action(a)} |")

    # 7. Vendor payments pending (native INR)
    L.append(f"\n## 7. Vendor Payments Pending — {vendor['count']} · {fmt_inr_full(vendor['total_inr'])}\n")
    if not vendor["connected"]:
        L.append("_No vendor payments sheet connected._")
    elif not vendor["pending"]:
        L.append("None pending. ✅")
    else:
        L.append("| Vendor | Invoice | Amount (INR) | Due | Overdue | Status |")
        L.append("|--------|---------|--------------|-----|---------|--------|")
        for v in vendor["pending"]:
            due = v["due"].strftime('%d %b %Y') if v["due"] else "—"
            L.append(f"| {v['vendor']} | {v['invoice'] or '—'} | {fmt_inr_full(v['amount_inr'])} "
                     f"| {due} | {v['overdue'] or '—'} | {v['status'] or '—'} |")

    # 8. Software & licenses inventory + renewals
    inv = get_software_inventory(data)
    renewals = get_upcoming_renewals(data, 60)
    L.append(f"\n## 8. Software & Licenses — {len(inv)} subscriptions\n")
    if not inv:
        L.append("_No software/subscription rows found._")
    else:
        L.append("| Application / Licence | Cost (this month) | Frequency | Owner/Dept | Renews |")
        L.append("|---|---|---|---|---|")
        for i in inv:
            L.append(f"| {i['app']} | {fmt_usd(i['cost']) if i['cost'] is not None else '—'} "
                     f"| {i['frequency'] or '—'} | {i['dept'] or '—'} "
                     f"| {i['renewal'].strftime('%d %b %Y') if i['renewal'] else '—'} |")
        L.append(f"\n**Renewals due in the next 60 days — {len(renewals)}**")
        if renewals:
            for r in renewals:
                L.append(f"- {r['renewal'].strftime('%d %b %Y')}: *{r['app']}* "
                         f"({fmt_usd(r['cost']) if r['cost'] is not None else '—'})")
        else:
            L.append("- None.")

    # 9. IT budget vs actual
    bva = get_budget_vs_actual(data)
    L.append("\n## 9. Laptop Procurement: Plan vs Actual\n")
    L.append("_'Planned per month' is the procurement-plan annual total ÷ 12, not a separately-approved budget._\n")
    L.append("| Item | Amount |")
    L.append("|---|---|")
    L.append(f"| Planned laptop procurement — annual FY26 (procurement plan) | {fmt_inr_full(bva['laptop_annual_inr'])} |")
    L.append(f"| Planned laptop procurement — per month (annual ÷ 12) | {fmt_inr_full(bva['laptop_monthly_inr'])} |")
    L.append(f"| Laptop spend — this month | {fmt_inr_full(bva['laptop_spend_inr'])}"
             + (f", {bva['laptop_pct_of_monthly']:.0f}% of planned monthly"
                if bva['laptop_pct_of_monthly'] is not None else "") + " |")
    laptop_models = ", ".join(f"{m['model']} ×{m['units']}" for m in bva["laptop_models"])
    L.append(f"| Laptops procured — this month | {bva['laptops_this_month']}"
             + (f" ({laptop_models})" if laptop_models else "") + " |")
    L.append(f"| Software & licenses — this month | {fmt_usd(bva['software_this_month'])} |")
    if bva["purchases_this_month"]:
        L.append("\n**Laptops in the purchase register this month**\n")
        L.append("| Model | Configuration | Purchase date | Serial |")
        L.append("|---|---|---|---|")
        for p in bva["purchases_this_month"]:
            L.append(f"| {(p['brand'] + ' ' + p['model']).strip() or '—'} | {p['configuration'] or '—'} "
                     f"| {p['date'].strftime('%d %b %Y') if p['date'] else '—'} | {p['serial'] or '—'} |")

    hist = bva["spend_history"]
    if hist["months"]:
        L.append("\n**Laptop spend by month (FY26 to date)**\n")
        L.append("| Month | Laptops | Models | Spend (₹) |")
        L.append("|---|---|---|---|")
        for m in hist["months"]:
            mdl = ", ".join(f"{x['model']} ×{x['units']}" for x in m.get("models", [])) or "—"
            L.append(f"| {m['month']} | {m['units']} | {mdl} | {fmt_inr_full(m['spend_inr'])} |")
        L.append(f"| **YTD total** | **{hist['ytd_units']}** | | **{fmt_inr_full(hist['ytd_spend_inr'])}** |")

    # 10. Laptop delivery lead times by vendor
    options = get_delivery_options(data)
    if options:
        L.append("\n## 10. Laptop Delivery Lead Times\n")
        L.append("How quickly each vendor can deliver, per device. *Fastest* is the "
                 "quickest option (worst-case days).\n")
        L.append("| Device | For (depts) | Fastest | All vendors |")
        L.append("|---|---|---|---|")
        for o in options:
            fastest = (f"{o['fastest']['vendor']} ({o['fastest']['text']})"
                       if o['fastest'] else "—")
            allv = ", ".join(f"{v['vendor']}: {v['text']}" for v in o["vendors"]) or "—"
            L.append(f"| {o['device']} | {o['departments'] or '—'} | {fastest} | {allv} |")
        terms = get_payment_terms(data)
        if terms:
            L.append("\n**Vendor payment terms**\n")
            L.append("| Vendor | Payment terms |")
            L.append("|---|---|")
            for vendor, term in terms.items():
                L.append(f"| {vendor} | {term} |")

        # Combined view: vendors (fastest first) × device lead time + payment terms
        matrix = get_vendor_delivery_matrix(data)
        if matrix["rows"]:
            L.append("\n**By vendor — delivery + payment terms**\n")
            L.append(f"| Vendor | {' | '.join(matrix['devices'])} | Payment terms |")
            L.append("|" + "---|" * (len(matrix["devices"]) + 2))
            for r in matrix["rows"]:
                L.append("| " + " | ".join([r["vendor"], *r["cells"], r["terms"]]) + " |")

    L.append(f"\n---\n{_fx_footnote()}")
    L.append(f"\n_Generated: {TODAY.strftime('%d %B %Y')}_")
    return "\n".join(L)


def generate_weekly_full(data: dict, prev_snap: Optional[dict] = None) -> str:
    return build_report_full(data, prev_snap, "Weekly")


def generate_monthly_slack(data: dict, prev_snap: Optional[dict] = None) -> str:
    return build_report_slack(data, prev_snap, "Monthly")


def generate_monthly_full(data: dict, prev_snap: Optional[dict] = None) -> str:
    return build_report_full(data, prev_snap, "Monthly")


# ---------------------------------------------------------------------------
# Slack Block Kit rendering (richer visual layout)
# ---------------------------------------------------------------------------

def _blk_header(text: str) -> dict:
    return {"type": "header", "text": {"type": "plain_text", "text": text[:150], "emoji": True}}


def _blk_divider() -> dict:
    return {"type": "divider"}


def _blk_section(text: str) -> dict:
    if len(text) > 2900:
        text = text[:2860].rstrip() + "\n_…truncated — see full report_"
    return {"type": "section", "text": {"type": "mrkdwn", "text": text}}


def _blk_context(text: str) -> dict:
    return {"type": "context", "elements": [{"type": "mrkdwn", "text": text[:2900]}]}


def _blk_named_section(title: str, lines: list, empty: str) -> list:
    body = "\n".join(lines) if lines else empty
    return [_blk_divider(), _blk_section(f"{title}\n{body}")]


def build_report_blocks(data: dict, prev_snap: Optional[dict], period: str) -> list:
    """Render the seven-section report as Slack Block Kit blocks for a cleaner,
    dashboard-style post. Falls back to the text summary for notifications."""
    issues = get_it_issues(data)
    open_issues = [i for i in issues["issues"] if i["is_open"]]
    stock_os = get_stock_by_os(data)
    stock_ready = len(data["in_stock"])
    aging = get_aging_laptops(data)
    critical = [a for a in aging if a["priority"] == "Critical"]
    joiners = get_joiners_with_laptop_needs(data, 30)
    laptop_spend = get_laptop_spend(data)
    software_total = get_software_spend_this_month(data)
    runway = get_procurement_runway(data)
    vendor = get_vendor_payments(data)

    blocks = [_blk_header(f"📋 IT {period} Report — {TODAY.strftime('%d %b %Y')}")]

    # 1) Open tickets
    if not issues["connected"]:
        lines = ["_No ticket source connected_"]
    elif not open_issues:
        lines = ["None open ✅"]
    else:
        lines = [f"› *{_truncate(i['issue'], 55)}* — _{_truncate(i['remark'], 90)}_  "
                 f"`{i['status'] or 'Open'}` · raised by {i['raised_by'] or 'unknown'} "
                 f"· owner {i['owner'] or 'unassigned'}" for i in open_issues]
    blocks += _blk_named_section(f"*🐞 1) Open IT Tickets — {len(open_issues)}*", lines, "—")

    # 2) Stock ready by OS + configs
    if stock_ready == 0:
        lines = ["None in ready stock"]
    else:
        lines = []
        for os_label, items in stock_os.items():
            lines.append(f"*{os_label} — {len(items)}*")
            lines += [f"› {it['config']}" for it in items]
    blocks += _blk_named_section(f"*💻 2) Laptop Stock Ready — {stock_ready}*", lines, "—")

    # 3) Procurement suggestion
    need = len(joiners) + len(critical) - stock_ready
    if need > 0:
        line = (f"⚠️ Order *{need}* laptop(s) — {len(joiners)} joiners (30d) + "
                f"{len(critical)} critical replacements vs {stock_ready} ready")
    else:
        line = (f"✅ Stock covers demand — {stock_ready} ready vs {len(joiners)} joiners "
                f"+ {len(critical)} critical replacements")
    proc_lines = [line]
    if runway["weeks"] is not None:
        proc_lines.append(f"_Runway ~{runway['weeks']} wks at {runway['avg_per_week']} joiners/wk_")
    blocks += _blk_named_section("*🛒 3) Procurement Suggestion*", proc_lines, "—")

    # 4) Upcoming joiners
    if not joiners:
        lines = ["None in the next 30 days"]
    else:
        lines = []
        for j in joiners:
            days = f"in {j['days_until']}d" if j['days_until'] > 0 else "today"
            cfg = f" · _{j['laptop_config']}_" if j['laptop_config'] else ""
            lines.append(f"› {j['name']} — {j['department']} · DOJ {j['doj'].strftime('%d %b')} ({days}){cfg}")
    blocks += _blk_named_section(f"*⏰ 4) Upcoming Joiners (30d) — {len(joiners)}*", lines, "—")

    # 5) Spend — total software/license spend + laptop procurement
    blocks += _blk_named_section(
        f"*💰 5) Spend This Month — {_spend_month_label(data)}*",
        [f"*Apps & licenses* {fmt_usd(software_total)}",
         f"*Laptops* (procurement plan) {fmt_usd(laptop_spend['total_spend'])}"],
        "—")

    # 6) Aging + action
    if not aging:
        lines = ["None over 3.5 years ✅"]
    else:
        lines = [f"› {a['employee']} — {_short_model(a['make'], a['model'])} · "
                 f"{a['age_years']}yr → _{aging_action(a)}_" for a in aging[:8]]
        if len(aging) > 8:
            lines.append(f"_+{len(aging) - 8} more (oldest first)_")
    blocks += _blk_named_section(
        f"*⏳ 6) Laptop Aging — {len(aging)} over 3.5yr ({len(critical)} critical)*", lines, "—")

    # 7) Vendor payments
    if not vendor["connected"]:
        lines = ["_No vendor payments sheet connected_"]
    elif not vendor["pending"]:
        lines = ["None pending ✅"]
    else:
        lines = []
        for v in vendor["pending"]:
            due = v["due"].strftime('%d %b %Y') if v["due"] else "—"
            inv = f" ({v['invoice']})" if v["invoice"] else ""
            ov = f" · overdue {v['overdue']}" if v["overdue"] else ""
            lines.append(f"› {v['vendor']}{inv} — *{fmt_inr_full(v['amount_inr'])}* · due {due}{ov}")
    blocks += _blk_named_section(
        f"*🧾 7) Vendor Payments Pending — {vendor['count']} · {fmt_inr_full(vendor['total_inr'])}*", lines, "—")

    blocks.append(_blk_divider())
    blocks.append(_blk_context(f"{_fx_footnote()}  ·  React 👍/👎 or reply with feedback."))
    return blocks


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) < 2 or sys.argv[1] not in ("weekly", "monthly"):
        print("Usage: generate-report.py weekly|monthly", file=sys.stderr)
        sys.exit(1)

    report_type = sys.argv[1]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Loading Excel data from {DATA_DIR} …")
    data = load_data()
    truly_assigned = sum(1 for r in data['assigned'] if is_truly_assigned(r))
    print(f"  Loaded: {truly_assigned} assigned laptops ({len(data['assigned'])} rows, "
          f"{len(data['assigned']) - truly_assigned} excluded as In Stock/blank), "
          f"{len(data['history'])} history records, {len(data['spend'])} spend rows, "
          f"{len(data['joinings'])} joiners")

    # Load previous snapshot for week-over-week comparison
    prev_snap = load_previous_snapshot()
    if prev_snap:
        print(f"  Previous snapshot: {prev_snap.get('date', 'unknown')} "
              f"(stock_ready={prev_snap.get('stock_ready')}, "
              f"stock_backup={prev_snap.get('stock_backup')})")
    else:
        print("  No previous snapshot found — this is the first run.")

    if report_type == "weekly":
        slack = generate_weekly_slack(data, prev_snap)
        full = generate_weekly_full(data, prev_snap)
    else:
        slack = generate_monthly_slack(data, prev_snap)
        full = generate_monthly_full(data, prev_snap)

    (OUTPUT_DIR / "slack-summary.md").write_text(slack, encoding="utf-8")
    (OUTPUT_DIR / "full-report.md").write_text(full, encoding="utf-8")

    # Extra context for the IT Helper bot beyond the aggregate report: per-person
    # laptops, joiner readiness + onboarding, returns/offboarding, peripherals,
    # and open tickets with requester/age.
    (OUTPUT_DIR / "bot-context.md").write_text(
        build_bot_context(data), encoding="utf-8")

    # Slack Block Kit layout (richer visual post). post-to-slack.py uses these
    # blocks when present, with slack-summary.md as the notification fallback.
    blocks = build_report_blocks(data, prev_snap, "Weekly" if report_type == "weekly" else "Monthly")
    (OUTPUT_DIR / "slack-blocks.json").write_text(
        json.dumps(blocks, indent=2, default=str), encoding="utf-8")

    # Dump structured metrics so the docx generator and chart generator can
    # produce visuals without re-parsing markdown.
    hc = get_health_check(data, prev_snap)
    laptop_spend = get_laptop_spend(data)
    app_total, _, _, _ = get_current_month_spend(data)
    runway = get_procurement_runway(data)
    pace = get_spend_pace(data)
    aging_dist = get_age_distribution(data)
    aging_all = get_aging_laptops(data)
    joiners_7 = get_upcoming_joiners(data, 7)
    joiners_30 = get_upcoming_joiners(data, 30)
    total_assigned = sum(1 for r in data["assigned"] if is_truly_assigned(r))

    overall = "🟢 On Track"
    if "🔴" in (hc["stock"], hc["aging"], hc["joiner_prep"], hc["spend"]):
        overall = "🔴 Action Needed"
    elif "🟡" in (hc["stock"], hc["aging"], hc["joiner_prep"], hc["spend"]):
        overall = "🟡 Attention"

    readiness_pct = get_onboarding_readiness(data, 7)
    cost_per_joiner = get_cost_per_joiner(data)  # USD
    issues = get_it_issues(data)
    action_items = get_action_items(data, hc)

    metrics = {
        "report_type": report_type,
        "date": TODAY.isoformat(),
        "overall_status": overall,
        "health": {k: hc[k] for k in ("stock", "aging", "joiner_prep", "spend")},
        "kpis": {
            "total_laptops": total_assigned + len(data["in_stock"]) + len(data["backup"]),
            "total_assigned": total_assigned,
            "stock_ready": len(data["in_stock"]),
            "stock_backup": len(data["backup"]),
            "aging_total": len(aging_all),
            "aging_critical": hc["_critical_aging"],
            "joiners_next_7": hc["_joiners_7"],
            "joiners_next_30": len(joiners_30),
            "runway_weeks": runway["weeks"],
            "avg_assignments_per_week": runway["avg_per_week"],
            "laptop_spend_month": laptop_spend["total_spend"],  # USD
            "laptop_spend_pct_of_budget": pace["pct_used"],
            "monthly_budget_usd": pace["monthly_planned"],  # USD
            "app_spend_month_usd": app_total,
            "onboarding_readiness_pct": readiness_pct,
            "cost_per_joiner_usd": cost_per_joiner,
            "it_issues_open": issues["open"] if issues["connected"] else None,
            "it_issues_high_open": issues["high_open"] if issues["connected"] else None,
        },
        "aging_distribution": aging_dist,
        "top_aging_critical": [
            {"name": a["employee"], "model": f"{a['make']} {a['model']}".strip(),
             "age_years": a["age_years"], "remark": a["remark"]}
            for a in aging_all[:10]
        ],
        "stock_vs_demand": {
            "stock_ready": len(data["in_stock"]),
            "stock_backup": len(data["backup"]),
            "joiners_7d": hc["_joiners_7"],
            "joiners_30d": len(joiners_30),
        },
        "onboarding_pipeline": {
            "d7": len(get_upcoming_joiners(data, 7)),
            "d14": len(get_upcoming_joiners(data, 14)),
            "d30": len(joiners_30),
            "d90": len(get_upcoming_joiners(data, 90)),
        },
        "it_issues": {
            "connected": issues["connected"],
            "open": issues["open"],
            "resolved": issues["resolved"],
            "high_open": issues["high_open"],
        },
        "action_items": action_items,
        "risks": get_risk_callouts(data, hc),
    }
    (OUTPUT_DIR / "metrics.json").write_text(json.dumps(metrics, indent=2, default=str), encoding="utf-8")

    # Save new snapshot for next run
    save_snapshot(current_snapshot(data))
    print(f"  Saved new snapshot to {SNAPSHOT_DIR / 'latest.json'}")

    print(f"Reports saved to {OUTPUT_DIR}/")
    print(f"  slack-summary.md: {len(slack)} chars")
    print(f"  full-report.md: {len(full)} chars")


if __name__ == "__main__":
    main()
