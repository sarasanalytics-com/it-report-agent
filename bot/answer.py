#!/usr/bin/env python3
"""Answer engine for the IT Slack bot.

Pulls live IT data by reusing the existing pipeline scripts (fetch-excel,
fetch-issues, generate-report), builds a grounded context from the generated
report, and asks Claude to answer the user's question using ONLY that data.

The data is cached and refreshed on a TTL so rapid-fire questions don't
re-download the spreadsheets every time.

CLI usage (for local testing once env + data are set up):
    python bot/answer.py "how many laptops are in stock?"
    python bot/answer.py --no-refresh "what did we spend on software in June?"
"""

from __future__ import annotations

import os
import re
import sys
import time
import logging
import json
import pathlib
import datetime as dt
import threading
import subprocess

import anthropic

log = logging.getLogger("it-bot.answer")

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "output"

# Most-complete report (monthly includes everything weekly does and more).
REPORT_TYPE = os.environ.get("BOT_REPORT_TYPE", "monthly")
# How long (seconds) to reuse fetched data before refreshing.
REFRESH_TTL = int(os.environ.get("DATA_REFRESH_TTL", "300"))
# Haiku is cheap/fast and plenty for grounded Q&A; override for higher quality.
MODEL = os.environ.get("ANTHROPIC_MODEL", "claude-haiku-4-5")
MAX_ANSWER_TOKENS = int(os.environ.get("BOT_MAX_TOKENS", "900"))

_client: anthropic.Anthropic | None = None
_lock = threading.Lock()
_last_refresh = 0.0
_context_cache = ""
_refreshing = False

SYSTEM_PROMPT = """You are a friendly IT helper in Slack for the Head of HR, \
who is NOT technical. You answer her questions about company laptops and IT in \
simple, everyday language.

Who you're talking to:
- The Head of HR. She cares most about: are new joiners going to have a laptop \
ready on day one, how many spare laptops we have, the status of any IT requests \
raised by employees, and roughly what we're spending. Frame answers around \
people and readiness, not technical detail.

How to answer:
- Use plain, warm English. NO technical jargon or abbreviations. Avoid terms \
like "asset tag", "OS", "MDM", "procurement", "SKU", "aging buckets". If you \
must mention something technical, explain it in one short phrase.
- FORMAT FOR SLACK, not Markdown — this matters: bold is a SINGLE asterisk \
*like this*; do NOT use double asterisks **like this** (Slack shows the literal \
asterisks). Italic is _like this_. Inline values can use `backticks`. Do NOT use \
Markdown headings (#), bullet markers like "- " or "* ", or [text](links) — Slack \
won't render them. Use "• " for bullet points.
- Give a COMPLETE, easy-to-understand answer the non-technical HR head can act \
on — don't reduce it to a bare one-liner. Include ALL the relevant facts the \
data has for her question. For example, for a laptop purchase give the count, \
the model(s), the cost (in ₹ and $), the purchase date, and the vendor if it's \
in the data. For a person give their laptop, age, warranty, and any peripherals.
- Lay it out clearly: a short lead sentence, then "• " bullets (or a code-block \
table for multi-column data) with each detail labelled. *Bold* the key figures.
- Be thorough with the FACTS, but don't pad with opinions, recommendations, \
projections, "what this means", or "would you like me to…" offers unless asked. \
Never attribute a recommendation to a person/team that isn't in the data (e.g. \
don't say "the IT team recommends…"). NEVER invent anything: if a specific \
detail isn't in the data, say it's not recorded rather than guessing or omitting \
it silently.
- When the answer compares several items across columns, present it as a TABLE \
inside a ``` triple-backtick code block ``` with aligned, space-padded columns \
and a header row. Slack does NOT render markdown pipe tables, so always use a \
code block — never bare '|' rows. Keep it SIMPLE and phone-friendly: at most ~4 \
columns, short one-word headers, and do NOT add running-total / cumulative \
columns (e.g. 'YTD so far' next to each row) unless she asks — put any total in a \
single final row instead. One small table per topic; split topics into separate \
tables rather than one wide one.
- For a new-joiner question, give a clear *Yes / Not yet* on whether a laptop \
is ready, plus the joining date.
- For a person question ("what laptop does X have", "is X's laptop old?"), use \
the EMPLOYEE LAPTOP DIRECTORY — give their laptop, its age, and whether it's \
due for replacement. If the name isn't an exact match, say who you think they \
mean or ask her to confirm the spelling.
- For accessories ("does X have a monitor/headset?"), use PERIPHERALS BY PERSON.
- For "which laptops were replacements vs new joiners?", "was <name>'s laptop a \
new-joiner setup or a replacement?", or "which laptop (serial) went to whom and \
was it a replacement?", use the LAPTOP ASSIGNMENTS BY SERIAL section (it matches \
each assigned laptop's serial number to the Asset History tab). State the person, \
their laptop, and whether it was a *new joiner* or *replacement*; if a laptop has \
no Asset-History record, say the type isn't recorded rather than guessing.
- For onboarding ("is X ready?", "what's pending for the new joiner?"), use \
UPCOMING JOINERS — give the joining date, the laptop needed, and which checklist \
items are still pending; judge laptop readiness by comparing the laptop needed \
against spare stock.
- The weekly report shows ACTION items; when she wants more detail on one, \
answer from the report data:
  • Procurement ("why do we need to buy laptops?", "how many should we order?", \
"do we have enough spare laptops?") — explain it from the numbers: spare laptops \
ready now vs the demand, which is upcoming joiners in the next 30 days PLUS \
laptops over 3.5 years that are due for replacement. Give the shortfall and what's \
driving it.
  • Vendor payments ("which payments are pending or overdue?", "what do we owe \
suppliers?", "what's overdue?") — use the Vendor Payments Pending list: each \
vendor, the amount in ₹, the due date, and days overdue; lead with the total and \
the most-overdue one.
  • Aging / replacements ("which laptops need replacing?", "which are the oldest \
or most critical?", "whose laptop is too old?") — use the Laptop Aging list: the \
person, their laptop model, its age in years, and the recommended action, oldest \
first.
- For offboarding ("how many laptops are yet to be returned?", "who hasn't \
returned their laptop?"), use LAPTOPS YET TO RETURN; for "did X return their \
laptop?" use LAPTOP RETURNS — completed.
- For her team's requests ("any IT tickets for X?", "how long pending?"), use \
IT TICKETS — match the person in 'Raised by', and quote the days open.
- For software/subscriptions ("what apps do we pay for?", "what renews next \
month?", "are we within budget?") and laptop procurement ("how many laptops did \
we buy this month — which model, how much, when?"), use SOFTWARE & LICENSES and \
IT BUDGET vs ACTUAL. For laptops give the count, the model(s), the spend in ₹ \
(and $), and the purchase date(s) from the register if listed; if a purchase \
date isn't recorded, say so rather than guessing. For "which vendor did we buy \
from?" use the "Laptops by vendor" breakdown in IT BUDGET vs ACTUAL (it lists \
each vendor and how many laptops — and which models — came from them); only say \
the vendor isn't recorded if that breakdown is genuinely absent. For "how many \
laptops did we buy from <vendor>?" — ESPECIALLY when she also wants the serial \
numbers, who each is assigned to, and whether it's a replacement or new joiner — \
use the LAPTOP PURCHASES BY VENDOR — serial · assignee · type section. Lead with \
the count for that vendor, then show ONE code-block table with columns Asset ID, \
Serial, Laptop, Assigned to, New joiner/Replacement, and Purchased (date), listing \
EVERY laptop from that vendor (this is the one case where a long table is expected \
— do not truncate it). \
Use 'In stock / unassigned' where there's no assignee and 'not recorded' where \
there's no replacement/new-joiner record; never guess either. When she asks \
about budget vs actual and spend is BELOW the planned budget, always state how \
much was SAVED (use the "SAVED so far" figure in the data — do not compute it \
yourself); if spend is above budget, state the overspend amount instead.
- For "unplanned spends", "ad-hoc / off-budget purchases", or "what did we spend \
outside the budget?", use the UNPLANNED / AD-HOC SPENDS section and show it as \
its OWN separate table (do not merge it with the planned laptop budget). Include \
the total if one is given. If that section says no unplanned sheet is connected, \
say so plainly rather than guessing.
- For "have we sold any laptops?", "which laptops were disposed/retired/sold?", \
use the LAPTOPS SOLD / DISPOSED section (these are company laptops we retired, \
not sales to customers); if it says none are recorded, say so plainly. For "how \
many spare <USB hub / adapter / cable / non-laptop item> do we have?", use the \
OTHER ASSETS IN STOCK section and quote the quantity.
- For "how many backup / old / standby laptops do we have?" or "which models are \
in the backup pool?", use the BACKUP / SPARE LAPTOPS section (these are the older \
3+ year spare machines, separate from the main ready stock). For "what laptop / \
spec does a <role or department> get as standard?", use the STANDARD LAPTOP \
CONFIGURATIONS section. If either section says none are recorded, say so plainly.
- For delivery/vendor questions ("laptop delivery timelines from vendors", "how \
fast can we get a laptop?", "which vendor is fastest?"), show the LAPTOP DELIVERY \
& PAYMENT TERMS table as ONE code-block table (vendors fastest-first, with the \
per-device lead times AND each vendor's payment terms). For "will the joiner's \
laptop arrive in time?", use the joiner's delivery note and flag timing risk \
against the joining date.
- For money, keep the currency symbols exactly as in the data (vendor bills and \
laptop procurement are in ₹; app/software subscriptions are in $) and don't \
convert them yourself.

Strict rules (accuracy is critical — she relies on this):
- Every number, name, date, and status in your reply MUST appear in the DATA \
section below. Do not calculate, estimate, infer, round, or extrapolate values \
that aren't explicitly given. If asked for something the data doesn't directly \
contain (e.g. an exact count that isn't listed), say warmly that you don't have \
that detail and suggest she check with the IT team — do NOT substitute a related \
or approximate figure.
- Quote figures exactly as written in the data, including the currency symbol \
(vendor bills & laptop procurement in ₹, software subscriptions in $); never convert them.
- Answer ONLY the topic the user asked about. Be complete WITHIN that topic, but \
do NOT bundle in other topics. If she asks about spend, show spend only — don't \
also add joiners, replacements, headcount, procurement suggestions, or stock. \
Only combine topics when she explicitly asks for several or for an "overview". \
One question = one topic.
- If she asks something unrelated to laptops/IT/people-IT-needs, gently say it's \
outside what you can see.
"""


def _client_singleton() -> anthropic.Anthropic:
    global _client
    if _client is None:
        # Reads ANTHROPIC_API_KEY from the environment.
        _client = anthropic.Anthropic()
    return _client


def _run_step(cmd: list[str], timeout: int = 240) -> bool:
    """Run a pipeline script; return True on success. Never raises."""
    try:
        subprocess.run(cmd, cwd=ROOT, env=os.environ.copy(),
                       check=True, capture_output=True, text=True, timeout=timeout)
        return True
    except Exception as exc:  # noqa: BLE001 - we want to keep going regardless
        stderr = getattr(exc, "stderr", "") or str(exc)
        log.warning("pipeline step failed: %s — %s", " ".join(cmd), str(stderr)[-400:])
        return False


def _build_context() -> str:
    # Use only the human-readable report + bot context (correctly labelled
    # currencies). metrics.json is intentionally excluded — it carries USD-only
    # convenience fields (e.g. monthly_budget_usd) that confused the answers.
    parts = []
    full = OUTPUT / "full-report.md"
    directory = OUTPUT / "bot-context.md"
    if full.exists():
        parts.append("# LATEST IT REPORT\n" + full.read_text(encoding="utf-8"))
    if directory.exists():
        parts.append(directory.read_text(encoding="utf-8"))
    return "\n\n".join(parts)


def _do_refresh() -> None:
    """Re-fetch data and rebuild the cached context. Serialized by _lock."""
    global _last_refresh, _context_cache
    with _lock:
        log.info("Refreshing IT data …")
        _run_step(["python", "scripts/fetch-excel.py"])      # spreadsheets (critical)
        _run_step(["python", "scripts/fetch-issues.py"])     # ClickUp tickets (best-effort)
        _run_step(["python", "scripts/generate-report.py", REPORT_TYPE])
        ctx = _build_context()
        if ctx:
            _context_cache = ctx
            _last_refresh = time.time()
            log.info("Data refreshed (%d chars of context).", len(ctx))
        elif not _context_cache:
            _context_cache = "(no IT data is currently available)"


def warm() -> None:
    """Blocking initial load — call once at startup."""
    _do_refresh()


def get_context() -> str:
    """Return the cached context immediately. If it's missing, load it
    synchronously (first call only); if it's stale, kick off a background
    refresh but still answer from the current cache — so questions stay fast."""
    global _refreshing
    if not _context_cache:
        _do_refresh()
    elif (time.time() - _last_refresh) >= REFRESH_TTL and not _refreshing:
        _refreshing = True

        def _bg():
            global _refreshing
            try:
                _do_refresh()
            finally:
                _refreshing = False

        threading.Thread(target=_bg, daemon=True).start()
    return _context_cache or "(no IT data is currently available)"


# Back-compat alias (app startup warms via this name too).
def refresh(force: bool = False) -> str:
    _do_refresh() if force else get_context()
    return _context_cache


def force_refresh() -> str:
    """Synchronously re-download the sheets and rebuild the context, blocking
    until done. Use right after a source sheet is edited so the next answer
    reflects the change. Returns the fresh context (or a 'no data' note)."""
    _do_refresh()
    return _context_cache or "(no IT data is currently available)"


def get_report_blocks() -> tuple[list | None, str]:
    """Ensure data is current, then return the Block Kit report + a text
    fallback for an on-demand 'send me the report' request."""
    get_context()  # generates slack-blocks.json / slack-summary.md if needed
    blocks = None
    bpath = OUTPUT / "slack-blocks.json"
    if bpath.exists():
        try:
            blocks = json.loads(bpath.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            blocks = None
    spath = OUTPUT / "slack-summary.md"
    summary = spath.read_text(encoding="utf-8")[:2900] if spath.exists() else "IT report"
    return blocks, summary


def _aligned_table(headers: list[str], rows: list[list[str]]) -> str:
    """A fixed-width, space-padded monospace table (renders aligned in a Slack
    code block). Built deterministically so every row lines up."""
    widths = [len(h) for h in headers]
    for r in rows:
        for i, c in enumerate(r):
            widths[i] = max(widths[i], len(c))

    def _line(cells):
        return "  ".join(c.ljust(widths[i]) for i, c in enumerate(cells)).rstrip()

    out = [_line(headers), _line(["-" * w for w in widths])]
    out.extend(_line(r) for r in rows)
    return "\n".join(out)


def _nice_date(s) -> str:
    if not s or str(s) in ("None", "—", ""):
        return "—"
    try:
        return dt.datetime.strptime(str(s)[:10], "%Y-%m-%d").strftime("%d %b %Y")
    except ValueError:
        return str(s)


# Vendor-purchase questions ("how many laptops from NGSS?") deserve the FULL
# list, perfectly aligned — too many rows for the LLM to relay within its token
# budget, and it won't column-align them. So answer these deterministically.
_VENDOR_INTENT_RE = re.compile(
    r"laptop|purchas|bought|\bbuy\b|buying|procur|serial|how many|assigned", re.I)


_VENDOR_HEADERS = ["Asset ID", "Serial", "Laptop", "Assigned to", "Type", "Purchased"]


def _match_vendor(question: str) -> dict | None:
    """Return the purchase record for the known vendor named in the question
    (with laptops), if it reads like a 'laptops from <vendor>' ask. Else None."""
    path = OUTPUT / "vendor-purchases.json"
    if not path.exists():
        return None
    try:
        vp = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    vendors = vp.get("vendors") or []
    if not vendors or not _VENDOR_INTENT_RE.search(question):
        return None
    # Longest name first so "NGSS India" wins over "NGSS".
    for v in sorted(vendors, key=lambda x: -len(str(x.get("vendor") or ""))):
        name = str(v.get("vendor") or "").strip()
        if not name or name.lower() == "not recorded":
            continue
        if re.search(r"\b" + re.escape(name.lower()) + r"\b", question.lower()):
            return v
    return None


def _vendor_rows(match: dict) -> list[list[str]]:
    rows = []
    for it in match.get("laptops", []):
        rows.append([
            str(it.get("asset_id") or "—"),
            str(it.get("serial") or "—"),
            str(it.get("laptop") or "—"),
            str(it.get("assignee") or "In stock / unassigned"),
            str(it.get("type") or "not recorded"),
            _nice_date(it.get("date")),
        ])
    return rows


def _vendor_purchase_table_answer(question: str) -> str | None:
    """The full aligned monospace table as text (fallback when files can't be
    uploaded, and for CLI use)."""
    match = _match_vendor(question)
    if not match:
        return None
    rows = _vendor_rows(match)
    if not rows:
        return None
    lead = (f"*{match['vendor']}* — *{match['count']} laptop(s)* purchased. Full list "
            f"with serial numbers, who each is assigned to, whether it's a new joiner "
            f"or replacement, and the purchase date:")
    return lead + "\n```\n" + _aligned_table(_VENDOR_HEADERS, rows) + "\n```"


def _write_vendor_xlsx(vendor: str, headers: list[str], rows: list[list[str]],
                       path: pathlib.Path) -> None:
    """Write the vendor laptop list as a real .xlsx (bold header, sized columns)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[^A-Za-z0-9 _-]", "", vendor)[:31] or "Laptops"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    head_fill = PatternFill("solid", fgColor="2C2D30")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
    for i, h in enumerate(headers, start=1):
        longest = max([len(str(h))] + [len(str(r[i - 1])) for r in rows] or [0])
        ws.column_dimensions[get_column_letter(i)].width = min(longest + 2, 42)
    ws.freeze_panes = "A2"
    wb.save(path)


def _render_vendor_png(title: str, headers: list[str], rows: list[list[str]],
                       path: pathlib.Path) -> None:
    """Render the vendor laptop list as a clean table image (identical on every
    device). Lazily imports matplotlib so bot startup stays fast."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    n = len(rows)
    fig_h = max(1.8, 0.34 * (n + 2))
    fig, ax = plt.subplots(figsize=(14, fig_h))
    ax.axis("off")
    ax.set_title(title, fontsize=14, fontweight="bold", loc="left", pad=14)
    tbl = ax.table(cellText=rows, colLabels=headers, loc="upper left", cellLoc="left")
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(9.5)
    tbl.scale(1, 1.35)
    for (r, c), cell in tbl.get_celld().items():
        cell.set_edgecolor("#d9d9d9")
        if r == 0:
            cell.set_facecolor("#2c2d30")
            cell.set_text_props(color="white", fontweight="bold")
        elif r % 2 == 0:
            cell.set_facecolor("#f4f4f5")
    try:
        tbl.auto_set_column_width(col=list(range(len(headers))))
    except Exception:  # noqa: BLE001 - layout best-effort
        pass
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def build_vendor_purchase_artifacts(question: str) -> dict | None:
    """For a 'laptops from <vendor>' question, build an Excel file + a table
    image and return {summary, files, text_table}. Returns None if the question
    isn't a vendor-purchase ask. Ensures data is fresh first."""
    get_context()  # refresh + (re)write vendor-purchases.json if stale
    match = _match_vendor(question)
    if not match:
        return None
    rows = _vendor_rows(match)
    if not rows:
        return None
    vendor, count = match["vendor"], match["count"]
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", vendor).strip("_")[:40] or "vendor"
    title = f"{vendor} — {count} laptop(s) purchased"
    files: list[str] = []
    xlsx = OUTPUT / f"laptops_from_{safe}.xlsx"
    try:
        _write_vendor_xlsx(vendor, _VENDOR_HEADERS, rows, xlsx)
        files.append(str(xlsx))
    except Exception:  # noqa: BLE001
        log.exception("vendor xlsx build failed")
    png = OUTPUT / f"laptops_from_{safe}.png"
    try:
        _render_vendor_png(title, _VENDOR_HEADERS, rows, png)
        files.append(str(png))
    except Exception:  # noqa: BLE001
        log.exception("vendor table image render failed")
    summary = (f"*{vendor}* — *{count} laptop(s)* purchased. Full list attached "
               f"(spreadsheet + image): serial numbers, who each is assigned to, "
               f"new joiner vs replacement, and the purchase date.")
    return {"summary": summary, "files": files,
            "text_table": _vendor_purchase_table_answer(question),
            "vendor": vendor, "count": count}


def answer_question(question: str) -> str:
    """Answer a natural-language question from the live IT data."""
    context = get_context()
    # Deterministic full-table answer for "laptops from <vendor>" questions.
    direct = _vendor_purchase_table_answer(question)
    if direct is not None:
        return direct
    client = _client_singleton()
    msg = client.messages.create(
        model=MODEL,
        max_tokens=MAX_ANSWER_TOKENS,
        system=SYSTEM_PROMPT + "\n\n===== DATA =====\n" + context[:150000],
        messages=[{"role": "user", "content": question}],
    )
    text = "".join(b.text for b in msg.content if getattr(b, "type", "") == "text").strip()
    return text or "I couldn't find an answer to that in the current IT data."


def _main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
    args = sys.argv[1:]
    force = True
    if "--no-refresh" in args:
        args.remove("--no-refresh")
        # Use whatever context already exists without re-fetching.
        global _context_cache, _last_refresh
        _context_cache = _build_context() or "(no IT data is currently available)"
        _last_refresh = time.time()
        force = False
    if not args:
        print('Usage: python bot/answer.py [--no-refresh] "your question"', file=sys.stderr)
        sys.exit(1)
    question = " ".join(args)
    if force:
        refresh(force=True)
    print(answer_question(question))


if __name__ == "__main__":
    _main()
